# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import json

import openpyxl
from openpyxl.cell import MergedCell

teacher_row_start = 10


# teachers start at row 10, col B - E
# subjects start at col J, row 6
# class is in col J, row 3
#  subjects from other classes are separated by empty cells or start with (...)
# required lessons in UE for subjects are in row 9 (below subjects)


def get_excel_rect_data_as_array(ws, min_row, max_row, min_col, max_col):
    array = []
    for row in range(min_row, max_row + 1):
        array.append([])
        all_none = True
        for col in range(min_col, max_col + 1):
            data = ws.cell(row=row, column=col).value
            if data is not None:
                all_none = False
            array[-1].append(data)

        if all_none:
            # remove last row
            del array[-1]
            break

    return array


def get_cell_range_from_merged_cells(ws, cell):
    merged_cells = ws.merged_cells.ranges
    for merged_cell in merged_cells:
        if cell.coordinate in merged_cell:
            return merged_cell.bounds  # row, col, row, col


def extract_classes_with_data_from_sheet(ws):
    start_row = 3
    start_col = 10  # J
    subjects_start_row = 6
    subjects_hours_year = 7
    subjects_hours_term = 9  # for one half year (one term/semester)

    all_classes = []
    curr_class = ws.cell(row=start_row, column=start_col)  # J3

    # some classes have 1 / 2 / 3 rows ...
    # the subjects for all classes start in row 6
    while curr_class.value is not None:
        range_class_name = get_cell_range_from_merged_cells(ws, curr_class)
        class_name_lines = []

        class_name_lines.append(curr_class.value)
        # extract class data
        if range_class_name[1] == range_class_name[3]:
            # whole class in one merged cell (same row)
            # so check the next rows until we get to row 6
            class_data_2 = ws.cell(row=range_class_name[1] + 1, column=range_class_name[0])
            range_class_data_2 = get_cell_range_from_merged_cells(ws, class_data_2)
            class_name_lines.append(class_data_2.value)

            if range_class_data_2[1] == range_class_data_2[3]:
                # one row
                class_data_3 = ws.cell(row=range_class_data_2[1] + 1, column=range_class_data_2[0])
                range_class_data_3 = get_cell_range_from_merged_cells(ws, class_data_3)
                if class_data_3.value is None:
                    class_name_lines.append("")
                else:
                    class_name_lines.append(class_data_3.value)

            else:
                # data is a merged cell with more than one row -> there is no data 3
                pass
        else:
            # data 1 is more than one row high
            if range_class_name[3] >= subjects_start_row:
                # we only have one data in a big merged cell
                raise Exception("TODO")
            else:
                # data 1 is 2 rows high, so we can have data 3
                class_data_3 = ws.cell(row=range_class_name[1] + 1, column=range_class_name[0])
                range_class_data_3 = get_cell_range_from_merged_cells(ws, class_data_3)
                if class_data_3.value is None:
                    class_name_lines.append("")
                else:
                    class_name_lines.append(class_data_3.value)

        class_obj = {
            "name": str.join("\n", class_name_lines),
            "name_fields": class_name_lines,
            "col_range": [range_class_name[0], range_class_name[2]],
            "subjects": [],  # {"name", "col", "coord"}
            "fake_subjects": []
        }
        all_classes.append(class_obj)
        curr_class = ws.cell(row=start_row, column=range_class_name[2] + 1)

    # now we have all classes
    # extract subjects for each class
    for class_obj in all_classes:
        class_col_range = class_obj["col_range"]
        for col in range(class_col_range[0], class_col_range[1] + 1):
            subject_cell = ws.cell(row=subjects_start_row, column=col)

            # some misc cell, e.g. sum of hours or something
            if subject_cell.value is None:
                continue

            # we have fake subjects e.g. managing a class
            subject_hours_total = ws.cell(row=subjects_hours_year, column=col)
            subject_hours_term = ws.cell(row=subjects_hours_term, column=col)

            subject_obj = {
                "name": subject_cell.value,
                "col": subject_cell.column,
                "coord": subject_cell.coordinate,
                "hours_total": subject_hours_total.value,
                "hours_term": subject_hours_term.value,
                "teachers_with_hours": []  # {"teacher_key", "hours"}
            }

            if subject_hours_term.value is None:
                # not a real subject...
                class_obj["fake_subjects"].append(subject_obj)
            else:
                class_obj["subjects"].append(subject_obj)

    return all_classes


def extract_and_set_teacher_hours_from_sheet(ws, all_classes, all_teachers_list):
    for class_obj in all_classes:
        subject_objs = class_obj["subjects"]
        for subject_obj in subject_objs:
            subject_col = subject_obj["col"]

            for row_index, teacher_obj in enumerate(all_teachers_list):
                hours_for_teacher_cell = ws.cell(row=teacher_row_start + row_index, column=subject_col)

                if hours_for_teacher_cell.value is None:
                    continue

                subject_obj["teachers_with_hours"].append({
                    "teacher_key": teacher_obj["key"],
                    "hours": hours_for_teacher_cell.value
                })


def get_all_teachers_from_rect_data(teacher_datas):
    all_teachers_dict = {}
    all_teachers_list = []

    for teacher_data in teacher_datas:
        i = 0
        contract_form = teacher_data[i]
        i += 1
        last_name = teacher_data[i]
        i += 1
        first_name = teacher_data[i]
        i += 1
        teacher_key = teacher_data[i]
        teacher_obj = {
            "contract_form": contract_form,
            "last_name": last_name,
            "first_name": first_name,
            "key": teacher_key
        }
        all_teachers_dict[teacher_key] = teacher_obj
        all_teachers_list.append(teacher_obj)

    return all_teachers_list, all_teachers_dict


def extract_all_data_from_sheet(ws):
    # key is "teacher key" (or nummer in excel/german)
    # value is hash with data
    teachers = dict()

    teacher_end_row = 100
    teacher_start_col = 2
    teacher_end_col = 5
    teacher_datas = get_excel_rect_data_as_array(ws, teacher_row_start, teacher_end_row, teacher_start_col,
                                                 teacher_end_col)
    all_teachers_list, all_teachers_dict = get_all_teachers_from_rect_data(teacher_datas)
    all_classes = extract_classes_with_data_from_sheet(ws)
    extract_and_set_teacher_hours_from_sheet(ws, all_classes, all_teachers_list)

    return {
        "all_teachers_list": all_teachers_list,
        "all_teachers_dict": all_teachers_dict,
        "all_classes": all_classes,
    }


def extract_teacher_available_color_mapping(wb_prefs):
    # excel uses aarrggbb for colors, use fgColor in fill!! (bg is for patterns)
    ws_color_legend = wb_prefs["Tabelle1"]
    mapping_allowed_row = 23
    mapping_allowed_col = 2

    mapping_allowed_cell = ws_color_legend.cell(row=mapping_allowed_row, column=mapping_allowed_col)
    allowed_color = mapping_allowed_cell.fill.fgColor

    mapping_not_allowed_row = 24
    mapping_not_allowed_col = 2

    mapping_not_allowed_cell = ws_color_legend.cell(row=mapping_not_allowed_row, column=mapping_not_allowed_col)
    mapping_not_allowed_color = mapping_not_allowed_cell.fill.fgColor

    print(f"allowed color: {allowed_color}")
    print(f"not allowed color: {mapping_not_allowed_color}")

    return {
        "allowed_bg_color": allowed_color,
        "not_allowed_bg_color": mapping_not_allowed_color
    }


def extract_single_teacher_preferences_from_sheet(ws, teacher_obj, color_legend_teacher_availability):
    start_col = 3  # C, here is also the name of the class and time slots
    end_col = 7  # G

    start_row = 2  # first time slot
    end_row = 10  # last time slot

    date_cells = []

    # C2 - G2 are the dates
    for i in range(start_col, end_col + 1):
        date_cell = ws.cell(row=start_row, column=i)
        date_cells.append(date_cell.value)

    preferences_cells = []  # 2d, column wise

    has_known_color = None

    # real data
    for col_i in range(start_col, end_col + 1):
        day_slots = []
        for row_j in range(start_row + 1, end_row + 1):
            pref_cell = ws.cell(row=row_j, column=col_i)

            cell_obj = {
                "slot_index": pref_cell.row - (start_row + 1),
                "class_name": pref_cell.value,  # can be None if empty
                "color": pref_cell.fill.fgColor,
            }

            if cell_obj["color"] == color_legend_teacher_availability["not_allowed_bg_color"] or cell_obj["color"] == \
                    color_legend_teacher_availability["allowed_bg_color"]:
                if has_known_color is not None:
                    if cell_obj["color"] != has_known_color:
                        print(f"warning: teacher {teacher_obj['key']} has both allowed and not allowed color")

                if cell_obj["color"] == color_legend_teacher_availability["not_allowed_bg_color"]:
                    has_known_color = cell_obj["color"]
                if cell_obj["color"] == color_legend_teacher_availability["allowed_bg_color"]:
                    has_known_color = cell_obj["color"]

            day_slots.append(cell_obj)
        preferences_cells.append(day_slots)

    pass


def extract_subject_mapping_from_sheet(ws_dozenten, all_teachers_dict):
    start_col = 1
    max_col = 100
    # no end col, we read until we find a class without a trailing dot
    start_row = 1
    max_row = 100

    subject_teacher_pair_by_class_name_dict = {}

    for col_i in range(start_col, max_col):
        class_name_cell = ws_dozenten.cell(row=start_row, column=col_i)
        class_name = class_name_cell.value

        subject_teacher_pair = []
        subject_teacher_pair_by_class_name_dict[class_name] = subject_teacher_pair

        for row_j in range(start_row + 1, max_row):
            subject_teacher_pair_cell = ws_dozenten.cell(row=row_j, column=col_i)
            value = subject_teacher_pair_cell.value

            # value is a pair, e.g. Eth(Bor) -> Eth, Bor
            if value is not None:
                # print(f"{subject_teacher_pair_cell.coordinate} = {value}")

                # if the value does not contain "(", ignore
                if "(" not in value:
                    print(f"warning: subject teacher pair does not contain '(' for class {class_name}")
                    continue

                subject, teacher = value.split("(")
                teacher_key = teacher.strip(")")

                if teacher_key == "":
                    print(f"warning: teacher key is empty for subject {subject} for class {class_name}")
                    continue
                if teacher_key not in all_teachers_dict:
                    print(f"warning: teacher key {teacher_key} not found for subject {subject} for class {class_name}")
                    continue

                subject_teacher_pair.append({"subject": subject, "teacher_key": teacher_key})

    return subject_teacher_pair_by_class_name_dict


# TODO rework
def extract_teachers_availability_and_prefs_from_sheet(wb_prefs, all_teachers_dict, color_legend_teacher_availability):
    dozenten_worksheet = wb_prefs["Dozenten"]

    # not needed anymore, we now use extract_subject_key_to_subject_mapping_from_sheet
    # mapping = extract_subject_mapping_from_sheet(dozenten_worksheet, all_teachers_dict)

    # each teacher has a separate sheet with it's key
    # tuple of worksheet and teacher key
    relevant_work_sheets = []

    for sheet in wb_prefs:
        worksheet_name = sheet.title  # to lower?

        if worksheet_name in all_teachers_dict:
            relevant_work_sheets.append((sheet, worksheet_name))

    for sheet, teacher_key in relevant_work_sheets:
        teacher_obj = all_teachers_dict[teacher_key]
        extract_single_teacher_preferences_from_sheet(sheet, teacher_obj, color_legend_teacher_availability)

    return None


# TODO
def extract_subject_key_to_subject_mapping_from_sheet(wb_prefs, all_subjects_dict):
    mapping_worksheet = wb_prefs["FächerMap"]

    subject_name_to_short_dict = {}

    # subject short | subject name
    start_col = 1
    start_row = 2
    max_row = 1000  # until we find empty row, but to be safe here

    for row_j in range(start_row, max_row):
        subject_short_cell = mapping_worksheet.cell(row=row_j, column=start_col)
        subject_name_cell = mapping_worksheet.cell(row=row_j, column=start_col + 1)

        if subject_short_cell.value is None:
            break

        subject_short = subject_short_cell.value
        subject_name = subject_name_cell.value
        subject_name_to_short_dict[subject_short] = subject_name

    return subject_name_to_short_dict


def extract_class_mapping_from_sheet(wb_prefs):
    mapping_worksheet = wb_prefs["KlassenMap"]
    # mapping from class keys to class full names (3 columns)
    class_key_to_full_name_dict = {}

    # class key | class name part 1 | class name part 2 (optional) | class name part 3 (optional)
    start_col = 1
    start_row = 2
    max_row = 100  # until we find empty row, but to be safe here

    for row_j in range(start_row, max_row):
        class_key = mapping_worksheet.cell(row=row_j, column=start_col)
        class_name_part_1 = mapping_worksheet.cell(row=row_j, column=start_col + 1)
        class_name_part_2 = mapping_worksheet.cell(row=row_j, column=start_col + 2)
        class_name_part_3 = mapping_worksheet.cell(row=row_j, column=start_col + 3)

        if class_key.value is None:
            continue

        name_full = ""

        if class_name_part_1.value is not None:
            name_full += class_name_part_1.value
        else:
            raise Exception(f"class name part 1 is required in worksheet KlassenMap for class key '{class_key.value}'")

        if class_name_part_2.value is not None:
            name_full += f" {class_name_part_2.value}"

        if class_name_part_3.value is not None:
            name_full += f" {class_name_part_3.value}"

        class_key_to_full_name_dict[class_key.value] = {
            "name_part_1": class_name_part_1.value,
            "name_part_2": class_name_part_2.value,
            "name_part_3": class_name_part_3.value,
            "name_full": name_full
        }

    return class_key_to_full_name_dict


def add_class_key_to_all_classes(all_classes, class_key_to_full_name_dict):

    classes_to_remove = []

    for class_obj in all_classes:
        class_name_fields = class_obj["name_fields"]

        class_name_fields_not_empty = [x for x in class_name_fields if x is not None and x != ""]
        class_name_full = str.join(" ", class_name_fields_not_empty)

        for class_key, class_name_infos in class_key_to_full_name_dict.items():
            if class_name_infos['name_full'] == class_name_full:
                class_obj["key"] = class_key
                print(f"class '{class_name_full}' has key '{class_key}'")
                break
        if "key" not in class_obj:
            print(f"warning: class '{class_name_full}' has no key in sheet 'KlassenMap' but has data in hours data excel table, DISCARDING!")
            classes_to_remove.append(class_obj)

    for class_obj in classes_to_remove:
        all_classes.remove(class_obj)



# see extract_single_teacher_preferences_from_sheet
def extract_current_single_plan_from_sheet(ws_plan, curr_row, curr_col, class_name):
    curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)

    slots_per_day = 8

    dates = []
    # Mo till Fr (the dates)
    for col in range(curr_col + 1, curr_col + 5):
        date_cell = ws_plan.cell(row=curr_row + 1, column=col)
        dates.append(date_cell.value)

    table = []

    for col in range(curr_col + 1, curr_col + 5):
        day_slots = []
        for row in range(curr_row + 2, curr_row + slots_per_day + 2):
            cell = ws_plan.cell(row=row, column=col)
            day_slots.append(cell.value)

        table.append(day_slots)

    return dates, table


# extract the current state of the plan from the sheet
# the task is to fill out ONLY the missing fields
def extract_current_plan_from_sheet(wb_prefs):
    ws_plan = wb_prefs["Plan"]

    start_col = 2
    start_row = 10
    max_row = 1000

    # key is the class name
    current_plan_dict = {}

    col_increment = 8
    row_increment = 13

    curr_row = start_row
    curr_col = start_col

    all_table_data_dict = {}

    # we check vertically and then horizontally, then again vertically, ...
    is_finished_vertically = False
    is_finished_horizontally = False

    while is_finished_vertically == False:
        is_finished_horizontally = False

        while is_finished_horizontally == False:
            curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)
            class_key = curr_class_name_cell.value

            # TODO sometimes we have invalid class names -> check against dict if we know this class!!
            # e.g. Reserve1, ReserveX, ...

            if class_key is None:
                is_finished_horizontally = True
                break
            else:
                # extract single
                table_data = extract_current_single_plan_from_sheet(ws_plan, curr_row, curr_col, class_key)
                table_dates = table_data[0]
                table_column_data = table_data[1]

                all_table_data_dict[class_key] = [table_dates, table_column_data]

            curr_col += col_increment

        curr_col = start_col
        curr_row += row_increment
        curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)

        if curr_class_name_cell.value is None:
            is_finished_vertically = True
            break

    return all_table_data_dict


def _get_class_keys_to_all_classes_index_mapping(all_table_data_dict, all_classes):
    # all_table_data_dict[class_key] = [table_dates, table_column_data]

    used_class_keys = []

    for class_obj in all_classes:
        class_key = class_obj["key"]
        if class_key not in all_table_data_dict:
            print(f"warning: class key '{class_key}' not found in table data (Plan) but has data in hours data excel table")
            continue
        used_class_keys.append(class_key)
        class_obj["table_dates"] = all_table_data_dict[class_key]

    all_table_data_class_keys = list(all_table_data_dict.keys())

    difference_class_keys = list(set(all_table_data_class_keys) - set(used_class_keys))

    for class_key in difference_class_keys:
        print(f"warning: class key '{class_key}' has table data (Plan) but was not found in all classes (hours data excel table)")

    pass


def main():
    wb = openpyxl.load_workbook('example/SJ 25-26_Gesamtübersicht Einsatz Lehrkräfte EA Halle 2025-05-21_3.xlsx',
                                read_only=False)
    worksheet_blacklist = ["variablen", "kumuliert"]
    all_relevant_work_sheets = []

    for sheet in wb:
        worksheet_name = sheet.title.lower()
        is_relevant = True
        for blacklisted_name in worksheet_blacklist:
            if blacklisted_name in worksheet_name:
                is_relevant = False
                break

        if not is_relevant:
            continue
        all_relevant_work_sheets.append(sheet)
        print(sheet.title)

    all_teachers_list = []
    all_teachers_dict = {}
    all_classes = []

    # ws = wb['KP 25_26']
    for ws in all_relevant_work_sheets:
        data = extract_all_data_from_sheet(ws)
        for teacher_obj in data["all_teachers_list"]:
            teacher_key = teacher_obj["key"]
            if teacher_key not in all_teachers_dict:
                all_teachers_dict[teacher_key] = teacher_obj
                all_teachers_list.append(teacher_obj)
            else:
                pass

        for _class in data["all_classes"]:
            all_classes.append(_class)

        # break
    print(f"found {len(all_teachers_list)} teachers")
    print(f"found {len(all_classes)} classes")
    #
    # # write to json file
    # teacher_data_json = json.dumps(all_teachers_list, indent=4)
    # with open('example/all_teachers.json', 'w') as outfile:
    #     outfile.write(teacher_data_json)

    # read from json file
    # with open('example/all_teachers.json', 'r') as infile:
    #     all_teachers_list = json.load(infile)
    #
    #     for teacher_obj in all_teachers_list:
    #         teacher_key = teacher_obj["key"]
    #         all_teachers_dict[teacher_key] = teacher_obj

    wb_teachers_preferences = openpyxl.load_workbook('example/07_KW 45_03.11.-07.11.2025_IN.xlsm',
                                                     read_only=False, data_only=True)

    color_legend_teacher_availability = extract_teacher_available_color_mapping(wb_teachers_preferences)

    extract_teachers_availability_and_prefs_from_sheet(wb_teachers_preferences, all_teachers_dict,
                                                       color_legend_teacher_availability)

    class_key_to_full_name_dict = extract_class_mapping_from_sheet(wb_teachers_preferences)
    add_class_key_to_all_classes(all_classes, class_key_to_full_name_dict)

    extract_subject_key_to_subject_mapping_from_sheet(wb_teachers_preferences, None)

    all_table_data_dict = extract_current_plan_from_sheet(wb_teachers_preferences)

    _get_class_keys_to_all_classes_index_mapping(all_table_data_dict, all_classes)

    print("finished")


main()
