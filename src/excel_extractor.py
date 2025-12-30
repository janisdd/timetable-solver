# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import json

import openpyxl


# teachers start at row 10, col B - E
# subjects start at col J, row 6
# class is in col J, row 3
#  subjects from other classes are separated by empty cells or start with (...)
# required lessons in UE for subjects are in row 9 (below subjects)

class ExcelExtractor:
    def __init__(self, excel_file_all_classes_path, excel_file_plan_availability_path):
        self.excel_file_all_classes_path = excel_file_all_classes_path
        self.excel_file_plan_availability_path = excel_file_plan_availability_path

        self.color_legend_teacher_availability = None
        self.all_teachers_list = None
        self.all_teachers_dict = None
        self.all_classes = None

    def read_all(self):
        wb = openpyxl.load_workbook(self.excel_file_all_classes_path, read_only=False)
        worksheet_blacklist = ["variablen", "kumuliert"]
        all_relevant_work_sheets = []

        for sheet in wb:
            worksheet_name = sheet.title.lower()
            is_relevant = True
            for blacklisted_name in worksheet_blacklist:
                if blacklisted_name in worksheet_name:
                    is_relevant = False
                    break

            if not is_relevant:
                continue
            all_relevant_work_sheets.append(sheet)
            print(sheet.title)

        self.all_teachers_list = []
        self.all_teachers_dict = {}
        self.all_classes = []

        # ws = wb['KP 25_26']
        for ws in all_relevant_work_sheets:
            data = extract_all_data_from_sheet(ws)
            for teacher_obj in data["all_teachers_list"]:
                teacher_key = teacher_obj["key"]
                if teacher_key not in self.all_teachers_dict:
                    self.all_teachers_dict[teacher_key] = teacher_obj
                    self.all_teachers_list.append(teacher_obj)
                else:
                    pass

            for _class in data["all_classes"]:
                self.all_classes.append(_class)

            # break
        print(f"found {len(self.all_teachers_list)} teachers")
        print(f"found {len(self.all_classes)} classes")
        #
        # # write to json file
        # teacher_data_json = json.dumps(all_teachers_list, indent=4)
        # with open('example/all_teachers.json', 'w') as outfile:
        #     outfile.write(teacher_data_json)

        # read from json file
        # with open('example/all_teachers.json', 'r') as infile:
        #     all_teachers_list = json.load(infile)
        #
        #     for teacher_obj in all_teachers_list:
        #         teacher_key = teacher_obj["key"]
        #         all_teachers_dict[teacher_key] = teacher_obj

        wb_plan_preferences = openpyxl.load_workbook(self.excel_file_plan_availability_path, read_only=False,
                                                     data_only=True)

        self.color_legend_teacher_availability = extract_class_and_teacher_available_color_mapping(wb_plan_preferences)

        class_key_to_full_name_dict = extract_class_mapping_from_sheet(wb_plan_preferences)
        add_class_key_to_all_classes(self.all_classes, class_key_to_full_name_dict)

        # this is only for reading, no writing!
        extract_teachers_availability_and_prefs_from_sheet(wb_plan_preferences, self.all_teachers_dict,
                                                           self.color_legend_teacher_availability)

        _teacher_sanity_checks(self.all_teachers_dict, self.all_teachers_list, self.all_classes)

        make_teacher_availability_and_prefs_canonical(self.all_teachers_list, self.color_legend_teacher_availability)



        extract_subject_key_to_subject_mapping_from_sheet(wb_plan_preferences, None)

        all_table_data_dict = extract_current_plan_from_sheet(wb_plan_preferences,
                                                              self.color_legend_teacher_availability)


        # this also removes empty classes (no subjects)
        _validate_class_and_teachers(self.all_classes, self.all_teachers_list, self.all_teachers_dict)


        _get_class_keys_to_all_classes_index_mapping(all_table_data_dict, self.all_classes)

        # we are only allowed to write to "PLAN" sheet
        print("finished")

    def get_max_slots_per_day(self):
        first_teacher = self.all_teachers_list[0]
        max_slots_per_day = len(first_teacher['availability_preference_table'][0])
        return max_slots_per_day

    def get_max_days(self):
        return len(self.all_teachers_list[0]['availability_preference_table'])

def get_excel_rect_data_as_array(ws, min_row, max_row, min_col, max_col):
    array = []
    for row in range(min_row, max_row + 1):
        array.append([])
        all_none = True
        for col in range(min_col, max_col + 1):
            data = ws.cell(row=row, column=col).value
            if data is not None:
                all_none = False
            array[-1].append(data)

        if all_none:
            # remove last row
            del array[-1]
            break

    return array


def get_cell_range_from_merged_cells(ws, cell):
    merged_cells = ws.merged_cells.ranges
    for merged_cell in merged_cells:
        if cell.coordinate in merged_cell:
            return merged_cell.bounds  # row, col, row, col


def extract_classes_with_data_from_sheet(ws):
    start_row = 3
    start_col = 10  # J
    subjects_start_row = 6
    subjects_hours_year = 7
    subjects_hours_term = 9  # for one half year (one term/semester)

    all_classes = []
    curr_class = ws.cell(row=start_row, column=start_col)  # J3

    # some classes have 1 / 2 / 3 rows ...
    # the subjects for all classes start in row 6
    while curr_class.value is not None:
        range_class_name = get_cell_range_from_merged_cells(ws, curr_class)
        class_name_lines = []

        class_name_lines.append(curr_class.value)
        # extract class data
        if range_class_name[1] == range_class_name[3]:
            # whole class in one merged cell (same row)
            # so check the next rows until we get to row 6
            class_data_2 = ws.cell(row=range_class_name[1] + 1, column=range_class_name[0])
            range_class_data_2 = get_cell_range_from_merged_cells(ws, class_data_2)
            class_name_lines.append(class_data_2.value)

            if range_class_data_2[1] == range_class_data_2[3]:
                # one row
                class_data_3 = ws.cell(row=range_class_data_2[1] + 1, column=range_class_data_2[0])
                range_class_data_3 = get_cell_range_from_merged_cells(ws, class_data_3)
                if class_data_3.value is None:
                    class_name_lines.append("")
                else:
                    class_name_lines.append(class_data_3.value)

            else:
                # data is a merged cell with more than one row -> there is no data 3
                pass
        else:
            # data 1 is more than one row high
            if range_class_name[3] >= subjects_start_row:
                # we only have one data in a big merged cell
                raise Exception("TODO")
            else:
                # data 1 is 2 rows high, so we can have data 3
                class_data_3 = ws.cell(row=range_class_name[1] + 1, column=range_class_name[0])
                range_class_data_3 = get_cell_range_from_merged_cells(ws, class_data_3)
                if class_data_3.value is None:
                    class_name_lines.append("")
                else:
                    class_name_lines.append(class_data_3.value)

        class_obj = {
            "name": str.join("\n", class_name_lines),
            "name_fields": class_name_lines,
            "name_single_line": str.join(" ", class_name_lines),
            "col_range": [range_class_name[0], range_class_name[2]],
            "subjects": [],  # {"name", "col", "coord"}
            "fake_subjects": []
        }

        # TODO infinite loop...
        print("TODO infinite loop...")
        # _first_cell = ws.cell(row=range_class_name[0], column=range_class_name[1])
        # _last_cell = ws.cell(row=range_class_name[2], column=range_class_name[1])
        #
        # print(f"info: {class_obj['name_single_line']} has column range: {class_obj['col_range']} [{_first_cell.coordinate} - {_last_cell.coordinate}]")

        all_classes.append(class_obj)

        if class_obj['name_single_line'] == 'Erzieher ':
            print("TODO 'Erzieher ' ...")
            curr_class = ws.cell(row=start_row, column=range_class_name[2] + 2)
        else:
            curr_class = ws.cell(row=start_row, column=range_class_name[2] + 1)

    # now we have all classes
    # extract subjects for each class
    for class_obj in all_classes:
        class_col_range = class_obj["col_range"]
        for col in range(class_col_range[0], class_col_range[1] + 1):
            subject_cell = ws.cell(row=subjects_start_row, column=col)

            # some misc cell, e.g. sum of hours or something
            if subject_cell.value is None:
                continue

            # we have fake subjects e.g. managing a class
            subject_hours_total = ws.cell(row=subjects_hours_year, column=col)
            subject_hours_term = ws.cell(row=subjects_hours_term, column=col)

            subject_obj = {
                "name": subject_cell.value,
                "col": subject_cell.column,
                "coord": subject_cell.coordinate,
                "hours_total": subject_hours_total.value,
                "hours_term": subject_hours_term.value,
                "teachers_with_hours": []  # {"teacher_key", "hours"}
            }

            if subject_hours_term.value is None:
                # not a real subject...
                class_obj["fake_subjects"].append(subject_obj)
            else:
                class_obj["subjects"].append(subject_obj)

    return all_classes


def extract_and_set_teacher_hours_from_sheet(ws, all_classes, all_teachers_list, teacher_row_start):
    for class_obj in all_classes:
        subject_objs = class_obj["subjects"]
        for subject_obj in subject_objs:
            subject_col = subject_obj["col"]

            for row_index, teacher_obj in enumerate(all_teachers_list):
                hours_for_teacher_cell = ws.cell(row=teacher_row_start + row_index, column=subject_col)

                if hours_for_teacher_cell.value is None:
                    continue

                subject_obj["teachers_with_hours"].append({
                    "teacher_key": teacher_obj["key"],
                    "teacher_full_name": teacher_obj['teacher_full_name'],
                    "hours": hours_for_teacher_cell.value
                })


def get_all_teachers_from_rect_data(teacher_datas):
    all_teachers_dict = {}
    all_teachers_list = []

    for teacher_data in teacher_datas:
        i = 0
        contract_form = teacher_data[i]
        i += 1
        last_name = teacher_data[i]
        i += 1
        first_name = teacher_data[i]
        i += 1
        teacher_key = teacher_data[i]
        teacher_obj = {
            "contract_form": contract_form,
            "last_name": last_name,
            "first_name": first_name,
            "teacher_full_name" : f"{first_name} {last_name}",
            "key": teacher_key,
            "availability_preference_table": None
        }
        all_teachers_dict[teacher_key] = teacher_obj
        all_teachers_list.append(teacher_obj)

    return all_teachers_list, all_teachers_dict


def extract_all_data_from_sheet(ws):
    # key is "teacher key" (or nummer in excel/german)
    # value is hash with data
    teachers = dict()

    teacher_row_start = 10
    teacher_end_row = 100
    teacher_start_col = 2
    teacher_end_col = 5
    teacher_datas = get_excel_rect_data_as_array(ws, teacher_row_start, teacher_end_row, teacher_start_col,
                                                 teacher_end_col)
    all_teachers_list, all_teachers_dict = get_all_teachers_from_rect_data(teacher_datas)
    all_classes = extract_classes_with_data_from_sheet(ws)
    extract_and_set_teacher_hours_from_sheet(ws, all_classes, all_teachers_list, teacher_row_start)

    return {
        "all_teachers_list": all_teachers_list,
        "all_teachers_dict": all_teachers_dict,
        "all_classes": all_classes,
    }


def get_cell_color(cell):
    # excel uses #aarrggbb for colors, use fgColor in fill!! (bg is for patterns)
    return cell.fill.fgColor


def extract_class_and_teacher_available_color_mapping(wb_prefs):
    # excel uses aarrggbb for colors, use fgColor in fill!! (bg is for patterns)
    ws_color_legend = wb_prefs["Tabelle1"]
    mapping_allowed_row = 23
    mapping_allowed_col = 2

    # in individual teachers sheets
    mapping_allowed_cell = ws_color_legend.cell(row=mapping_allowed_row, column=mapping_allowed_col)
    allowed_color = get_cell_color(mapping_allowed_cell)

    mapping_not_allowed_row = 24
    mapping_not_allowed_col = 2

    # in individual teachers sheets
    mapping_not_allowed_cell = ws_color_legend.cell(row=mapping_not_allowed_row, column=mapping_not_allowed_col)
    mapping_not_allowed_color = get_cell_color(mapping_not_allowed_cell)

    mapping_do_not_process_row = 25
    mapping_do_not_process_col = 2

    # only in "Plan" sheet
    # this should be kept empty by us in "Plan" sheet
    mapping_do_not_process_cell = ws_color_legend.cell(row=mapping_do_not_process_row,
                                                       column=mapping_do_not_process_col)
    mapping_do_not_process_color = get_cell_color(mapping_do_not_process_cell)

    print(f"allowed color: {allowed_color}")
    print(f"not allowed color: {mapping_not_allowed_color}")
    print(f"do not process color: {mapping_do_not_process_color}")

    return {
        "allowed_bg_color": allowed_color,
        "not_allowed_bg_color": mapping_not_allowed_color,
        "do_not_process_bg_color": mapping_do_not_process_color
    }


def extract_and_set_single_teacher_availability_preferences_from_sheet(ws, teacher_obj,
                                                                       color_legend_teacher_availability):
    start_col = 3  # C, here is also the name of the class and time slots
    end_col = 7  # G

    start_row = 2  # first time slot
    end_row = 10  # last time slot

    date_cells = []

    # C2 - G2 are the dates
    for i in range(start_col, end_col + 1):
        date_cell = ws.cell(row=start_row, column=i)
        date_cells.append(date_cell.value)

    preferences_cells = []  # 2d, column wise

    has_known_color = None

    # real data
    for col_i in range(start_col, end_col + 1):
        day_slots = []
        for row_j in range(start_row + 1, end_row + 1):
            pref_cell = ws.cell(row=row_j, column=col_i)

            cell_obj = {
                "slot_index": pref_cell.row - (start_row + 1),
                "class_key": pref_cell.value,  # can be None if empty
                "color": get_cell_color(pref_cell),
                "allowed": None # will be set later (black/white list)
            }

            if cell_obj["color"] == color_legend_teacher_availability["not_allowed_bg_color"] or cell_obj["color"] == \
                    color_legend_teacher_availability["allowed_bg_color"]:
                if has_known_color is not None:
                    if cell_obj["color"] != has_known_color:
                        print(f"error: teacher {teacher_obj['key']} has both allowed and not allowed color")
                        return False

                if cell_obj["color"] == color_legend_teacher_availability["not_allowed_bg_color"]:
                    has_known_color = cell_obj["color"]
                if cell_obj["color"] == color_legend_teacher_availability["allowed_bg_color"]:
                    has_known_color = cell_obj["color"]

            day_slots.append(cell_obj)
        preferences_cells.append(day_slots)

    teacher_obj['availability_preference_table'] = preferences_cells

    return True


def extract_subject_mapping_from_sheet(ws_dozenten, all_teachers_dict):
    start_col = 1
    max_col = 100
    # no end col, we read until we find a class without a trailing dot
    start_row = 1
    max_row = 100

    subject_teacher_pair_by_class_name_dict = {}

    for col_i in range(start_col, max_col):
        class_name_cell = ws_dozenten.cell(row=start_row, column=col_i)
        class_name = class_name_cell.value

        subject_teacher_pair = []
        subject_teacher_pair_by_class_name_dict[class_name] = subject_teacher_pair

        for row_j in range(start_row + 1, max_row):
            subject_teacher_pair_cell = ws_dozenten.cell(row=row_j, column=col_i)
            value = subject_teacher_pair_cell.value

            # value is a pair, e.g. Eth(Bor) -> Eth, Bor
            if value is not None:
                # print(f"{subject_teacher_pair_cell.coordinate} = {value}")

                # if the value does not contain "(", ignore
                if "(" not in value:
                    print(f"warning: subject teacher pair does not contain '(' for class {class_name}")
                    continue

                subject, teacher = value.split("(")
                teacher_key = teacher.strip(")")

                if teacher_key == "":
                    print(f"warning: teacher key is empty for subject {subject} for class {class_name}")
                    continue
                if teacher_key not in all_teachers_dict:
                    print(f"warning: teacher key {teacher_key} not found for subject {subject} for class {class_name}")
                    continue

                subject_teacher_pair.append({"subject": subject, "teacher_key": teacher_key})

    return subject_teacher_pair_by_class_name_dict


# TODO rework
def extract_teachers_availability_and_prefs_from_sheet(wb_prefs, all_teachers_dict, color_legend_teacher_availability):
    # dozenten_worksheet = wb_prefs["Dozenten"]

    # not needed anymore, we now use extract_subject_key_to_subject_mapping_from_sheet
    # mapping = extract_subject_mapping_from_sheet(dozenten_worksheet, all_teachers_dict)

    # each teacher has a separate sheet with it's key
    # tuple of worksheet and teacher key
    relevant_work_sheets = []

    for sheet in wb_prefs:
        worksheet_name = sheet.title  # to lower?

        if worksheet_name in all_teachers_dict:
            relevant_work_sheets.append((sheet, worksheet_name))

    error_count = 0

    for sheet, teacher_key in relevant_work_sheets:
        teacher_obj = all_teachers_dict[teacher_key]
        all_ok = extract_and_set_single_teacher_availability_preferences_from_sheet(sheet, teacher_obj,
                                                                                    color_legend_teacher_availability)

        if not all_ok:
            error_count += 1

    if error_count > 0:
        raise Exception(f" {error_count} errors in teacher availability preferences")

    return None


def _teacher_sanity_checks(all_teachers_dict, all_teachers_list, all_classes):

    teachers_to_ignore = []
    # check every valid teacher has availability preferences
    for teacher_obj in all_teachers_list:
        key = teacher_obj["key"]
        availability_preference_table = teacher_obj['availability_preference_table']

        if availability_preference_table is None:
            print(f"warning: teacher '{key}' [{teacher_obj['teacher_full_name']}] [{teacher_obj['contract_form']}] has no availability preferences -> teacher will not be used!")
            teachers_to_ignore.append(teacher_obj)
            continue

        # check if teacher perf classes are known classes
        for day_index, availability_preference_list in enumerate(availability_preference_table):
            for slot_index, slot_obj in enumerate(availability_preference_list):
                prefilled_class = slot_obj['class_key']
                if prefilled_class is not None:
                    is_known_class = False
                    for class_obj in all_classes:
                        if prefilled_class == class_obj['key']:
                            is_known_class = True
                            break

                    if not is_known_class:
                        print(f"warning: teacher '{key}' [{teacher_obj['teacher_full_name']}] has pref for unknown class '{prefilled_class}' on day index '{day_index}' slot '{slot_index}' -> slot will be set to skip/not allowed and class removed because invalid")
                        slot_obj['allowed'] = False
                        slot_obj['class_key'] = None


    for teacher_obj in teachers_to_ignore:
        key = teacher_obj["key"]
        all_teachers_list.remove(teacher_obj)
        del all_teachers_dict[key]

    error_count = 0

    # make sure all teachers are still there in the class subject list
    for class_obj in all_classes:
        class_key = class_obj["key"]
        class_name_single_line = class_obj['name_single_line']

        for subject_obj in class_obj['subjects']:
            subject_name = subject_obj["name"]
            teachers_with_hours = subject_obj['teachers_with_hours']

            for teacher_info in teachers_with_hours:
                teacher_key = teacher_info['teacher_key']
                teacher_full_name = teacher_info['teacher_full_name']

                if teacher_key not in all_teachers_dict:
                    msg = f"ERROR: teacher '{teacher_key}' [{teacher_full_name}] not found in teacher list for class '{class_key}' [{class_name_single_line}] and subject '{subject_name}'"
                    print(msg)
                    # raise Exception(msg)
                    error_count += 1

    if error_count > 0:
        print(f"TODO {error_count} errors in teacher availability preferences")
    #     raise Exception(f" {error_count} errors in teacher availability preferences")


# we only use blacklists
def make_teacher_availability_and_prefs_canonical(all_teachers_list, color_legend_teacher_availability):

    for teacher_obj in all_teachers_list:
        key = teacher_obj["key"]
        availability_preference_table = teacher_obj['availability_preference_table']

        # no color -> blacklist
        # allow color -> invert all
        is_blacklist = True

        for day_slots_list in availability_preference_table:
            for slot_obj in day_slots_list:
                if slot_obj["color"] == color_legend_teacher_availability["allowed_bg_color"]:
                    is_blacklist = False
                    break
            if not is_blacklist:
                break

        if is_blacklist:
            # set allowed
            for day_slots_list in availability_preference_table:
                for slot_obj in day_slots_list:
                    prefilled_class = slot_obj["class_key"]
                    if prefilled_class is not None:
                        slot_obj["allowed"] = False # already filled -> don't change
                        continue

                    if slot_obj["color"] == color_legend_teacher_availability["not_allowed_bg_color"]:
                        slot_obj["allowed"] = False  # should not be filled -> teacher not available here
                    else:
                        slot_obj["allowed"] = True

        else:
            # whitelist ... set all to false except where we have color allowed_bg_color
            # if something is prefilled -> don't change -> not allowed
            for day_slots_list in availability_preference_table:
                for slot_obj in day_slots_list:
                    prefilled_class = slot_obj["class_key"]
                    if prefilled_class is not None:
                        slot_obj["allowed"] = False  # already filled -> don't change
                        continue

                    if slot_obj["color"] == color_legend_teacher_availability["allowed_bg_color"]:
                        slot_obj["allowed"] = True  # should not be filled -> teacher not available here
                    else:
                        slot_obj["allowed"] = False


# TODO
def extract_subject_key_to_subject_mapping_from_sheet(wb_prefs, all_subjects_dict):
    mapping_worksheet = wb_prefs["FächerMap"]

    subject_name_to_short_dict = {}

    # subject short | subject name
    start_col = 1
    start_row = 2
    max_row = 1000  # until we find empty row, but to be safe here

    for row_j in range(start_row, max_row):
        subject_short_cell = mapping_worksheet.cell(row=row_j, column=start_col)
        subject_name_cell = mapping_worksheet.cell(row=row_j, column=start_col + 1)

        if subject_short_cell.value is None:
            break

        subject_short = subject_short_cell.value
        subject_name = subject_name_cell.value
        subject_name_to_short_dict[subject_short] = subject_name

    return subject_name_to_short_dict


def extract_class_mapping_from_sheet(wb_prefs):
    mapping_worksheet = wb_prefs["KlassenMap"]
    # mapping from class keys to class full names (3 columns)
    class_key_to_full_name_dict = {}

    # class key | class name part 1 | class name part 2 (optional) | class name part 3 (optional)
    start_col = 1
    start_row = 2
    max_row = 100  # until we find empty row, but to be safe here

    for row_j in range(start_row, max_row):
        class_key = mapping_worksheet.cell(row=row_j, column=start_col)
        class_name_part_1 = mapping_worksheet.cell(row=row_j, column=start_col + 1)
        class_name_part_2 = mapping_worksheet.cell(row=row_j, column=start_col + 2)
        class_name_part_3 = mapping_worksheet.cell(row=row_j, column=start_col + 3)

        if class_key.value is None:
            continue

        name_full = ""

        if class_name_part_1.value is not None:
            name_full += class_name_part_1.value
        else:
            raise Exception(f"class name part 1 is required in worksheet KlassenMap for class key '{class_key.value}'")

        if class_name_part_2.value is not None:
            name_full += f" {class_name_part_2.value}"

        if class_name_part_3.value is not None:
            name_full += f" {class_name_part_3.value}"

        class_key_to_full_name_dict[class_key.value] = {
            "name_part_1": class_name_part_1.value,
            "name_part_2": class_name_part_2.value,
            "name_part_3": class_name_part_3.value,
            "name_full": name_full
        }

    return class_key_to_full_name_dict


def add_class_key_to_all_classes(all_classes, class_key_to_full_name_dict):
    classes_to_remove = []

    for class_obj in all_classes:
        class_name_fields = class_obj["name_fields"]

        class_name_fields_not_empty = [x for x in class_name_fields if x is not None and x != ""]
        class_name_full = str.join(" ", class_name_fields_not_empty)

        for class_key, class_name_infos in class_key_to_full_name_dict.items():
            if class_name_infos['name_full'] == class_name_full:
                class_obj["key"] = class_key
                print(f"class '{class_name_full}' has key '{class_key}'")
                break
        if "key" not in class_obj:
            print(
                f"warning: class '{class_name_full}' has no key in sheet 'KlassenMap' but has data in hours data excel table, DISCARDING!")
            classes_to_remove.append(class_obj)

    for class_obj in classes_to_remove:
        all_classes.remove(class_obj)


# see extract_single_teacher_preferences_from_sheet
def extract_current_single_plan_from_sheet(ws_plan, curr_row, curr_col, class_name, color_legend_teacher_availability):
    curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)

    slots_per_day = 8
    num_days = 5

    dates = []
    # Mo till Fr (the dates)
    for col in range(curr_col + 1, curr_col + 1 + num_days):
        date_cell = ws_plan.cell(row=curr_row + 1, column=col)
        dates.append(date_cell.value)

    table = []

    for col in range(curr_col + 1, curr_col + 5):
        day_slots = []
        for row in range(curr_row + 2, curr_row + slots_per_day + 2):
            cell = ws_plan.cell(row=row, column=col)

            cell_color = get_cell_color(cell)
            # just ignore anything in that cell -> we should not fill / process it
            should_ignore_cell = cell_color == color_legend_teacher_availability["do_not_process_bg_color"]

            day_slots.append({
                "entry": cell.value,
                "ignore": should_ignore_cell
            })

        table.append(day_slots)

    return dates, table


# extract the current state of the plan from the sheet
# the task is to fill out ONLY the missing fields
def extract_current_plan_from_sheet(wb_prefs, color_legend_teacher_availability):
    ws_plan = wb_prefs["Plan"]

    start_col = 2
    start_row = 10
    max_row = 1000

    # key is the class name
    current_plan_dict = {}

    col_increment = 8
    row_increment = 13

    curr_row = start_row
    curr_col = start_col

    all_table_data_dict = {}

    # we check vertically and then horizontally, then again vertically, ...
    is_finished_vertically = False
    is_finished_horizontally = False

    while is_finished_vertically == False:
        is_finished_horizontally = False

        while is_finished_horizontally == False:
            curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)
            class_key = curr_class_name_cell.value

            # TODO sometimes we have invalid class names -> check against dict if we know this class!!
            # e.g. Reserve1, ReserveX, ...

            if class_key is None:
                is_finished_horizontally = True
                break
            else:
                # extract single
                table_data = extract_current_single_plan_from_sheet(ws_plan, curr_row, curr_col, class_key,
                                                                    color_legend_teacher_availability)
                table_dates = table_data[0]
                table_column_data = table_data[1]

                all_table_data_dict[class_key] = [table_dates, table_column_data]

            curr_col += col_increment

        curr_col = start_col
        curr_row += row_increment
        curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)

        if curr_class_name_cell.value is None:
            is_finished_vertically = True
            break

    return all_table_data_dict


def _get_class_keys_to_all_classes_index_mapping(all_table_data_dict, all_classes):
    # all_table_data_dict[class_key] = [table_dates, table_column_data]

    used_class_keys = []

    for class_obj in all_classes:
        class_key = class_obj["key"]
        if class_key not in all_table_data_dict:
            print(
                f"warning: class key '{class_key}' was not found in table data (Plan) but has data in hours data excel table with subjects")
            continue
        used_class_keys.append(class_key)
        class_obj['table_dates'] = all_table_data_dict[class_key]

    all_table_data_class_keys = list(all_table_data_dict.keys())

    difference_class_keys = list(set(all_table_data_class_keys) - set(used_class_keys))

    for class_key in difference_class_keys:
        print(
            f"warning: class key '{class_key}' has table data (Plan) but was not found in all classes (hours data excel table with subjects) -> ignoring")

    pass


def _validate_class_and_teachers(all_classes, all_teachers_list, all_teachers_dict):
    print("--- validating class teachers...")

    error_count = 0

    classes_to_remove = []

    for class_obj in all_classes:
        class_subjects = class_obj["subjects"]

        if len(class_subjects) == 0:
            print(f"warning: class '{class_obj['name_full']}' has no subjects -> will not be processed")
            continue

        class_subjects_to_remove = []

        for class_subject in class_subjects:
            subject_teacher_objs = class_subject["teachers_with_hours"]
            hours_term = class_subject["hours_term"]

            if type(hours_term) == str:
                print(
                    f"warning: class '{class_obj['name_single_line']}' has subject '{class_subject['name']}' with hours term '{hours_term}' -> will not be processed")
                class_subjects_to_remove.append(class_subject)
                continue

            if hours_term <= 0:
                print(
                    f"warning: class '{class_obj['name_single_line']}' has subject '{class_subject['name']}' with hours term '{hours_term}' -> will not be processed")
                class_subjects_to_remove.append(class_subject)
                continue

            if len(subject_teacher_objs) == 0:
                print(
                    f"warning: subject '{class_subject['name']}' of class '{class_obj['name_single_line']}' has no teachers assigned --> will not be processed")
                class_subjects_to_remove.append(class_subject)
                # error_count += 1
                continue

            for subject_teacher in subject_teacher_objs:
                teacher_key = subject_teacher["teacher_key"]
                teacher_hours = subject_teacher["hours"]

                if teacher_key not in all_teachers_dict:
                    print(
                        f"error: subject '{class_subject['name']}' of class '{class_obj['name_single_line']}' has unknown teacher key '{teacher_key}' assigned")
                    class_subjects_to_remove.append(class_subject)
                    error_count += 1
                    continue

        for class_subject in class_subjects_to_remove:
            class_obj["subjects"].remove(class_subject)

        if len(class_obj["subjects"]) == 0:
            print(
                f"warning: class '{class_obj['name_single_line']}' has no subjects left after validation -> will not be processed")
            classes_to_remove.append(class_obj)
            continue

    for class_obj in classes_to_remove:
        all_classes.remove(class_obj)
        print(
            f"warning: class '{class_obj['name_single_line']}' has no subjects left after validation -> class will not be processed")

    used_teachers = set()
    teachers_to_remove = []

    for class_obj in all_classes:
        class_key = class_obj['key']
        subjects_info = class_obj["subjects"]

        for subject_info in subjects_info:
            subject_key = subject_info["name"]
            teachers_with_hours = subject_info["teachers_with_hours"]  # list

            for teachers_with_hour_tuple in teachers_with_hours:
                teacher_key = teachers_with_hour_tuple['teacher_key']
                total_required_hours = teachers_with_hour_tuple['hours']
                used_teachers.add(teacher_key)

    for teacher_obj in all_teachers_list:
        teacher_key = teacher_obj["key"]
        if teacher_key not in used_teachers:
            print(f"warning: teacher '{teacher_key}' [{teacher_obj['teacher_full_name']}] not used in any class -> teacher will be ignored")
            teachers_to_remove.append(teacher_obj)

    for teacher_obj in teachers_to_remove:
        teacher_key = teacher_obj["key"]
        all_teachers_list.remove(teacher_obj)
        del all_teachers_dict[teacher_key]


    if error_count > 0:
        print(f"TODO {error_count} errors in teacher availability preferences")
        # raise Exception(f" {error_count} errors in teacher availability preferences")

    pass


if __name__ == '__main__':
    excel_extractor = ExcelExtractor(
        'example/SJ 25-26_Gesamtübersicht Einsatz Lehrkräfte EA Halle 2025-05-21_3.xlsx',
        'example/07_KW 45_03.11.-07.11.2025_IN.xlsm'
    )
    excel_extractor.read_all()
    excel_extractor.get_max_slots_per_day()
