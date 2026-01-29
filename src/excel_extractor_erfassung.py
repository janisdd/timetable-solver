# https://openpyxl.readthedocs.io/en/stable/tutorial.html

import openpyxl


class ExcelExtractorStundenerfassung:
    def __init__(self, all_excel_files_stundenerfassung, year_index):
        self.all_excel_files_stundenerfassung = all_excel_files_stundenerfassung
        self.year_index = year_index
        # the soll index of the year
        # 5 = J
        self.year_columns = [7, 12, 17, 22]
        self.all_soll_data_dict = {}
        print(f"[INFO] using year index: {self.year_index} -> column (1-based) {self.year_columns[self.year_index]}")

    def read_all(self):

        for excel_file_stundenerfassung_path in self.all_excel_files_stundenerfassung:
            stundenerfassung_obj = self.read_single_stundenerfassung(excel_file_stundenerfassung_path, self.year_index)

            self.all_soll_data_dict[stundenerfassung_obj["class_key"]] = stundenerfassung_obj["soll_data_array"]

        # we are only allowed to write to "PLAN" sheet
        print("finished reading excel files")

    def read_single_stundenerfassung(self, excel_files_stundenerfassung_path, year_index):
        workbook = openpyxl.load_workbook(excel_files_stundenerfassung_path, read_only=False, data_only=True)
        ws = workbook["LehrerStundenübersicht"]

        soll_data_array = []

        lfd_nr_col = 1  # A just some number
        subject_col = 2 # B
        teacher_key_col = 3 # C
        name_col = 4  # D
        teacher_subject_pair_col = 5  # E

        start_row = 33
        soll_start_col = self.year_columns[year_index]
        curr_row = start_row

        class_key = ws.cell(row=1, column=1).value

        curr_nr_cell = ws.cell(row=curr_row, column=lfd_nr_col)

        while curr_nr_cell.value is not None:
            print(curr_nr_cell.value)
            lfd_nr_value = curr_nr_cell.value
            teacher_name_value = ws.cell(row=curr_row, column=name_col).value
            # abkürzung
            teacher_subject_pair_value = ws.cell(row=curr_row, column=teacher_subject_pair_col).value
            # soll x. jahr
            soll_value = ws.cell(row=curr_row, column=soll_start_col).value
            # ist akkum x. jahr
            ist_akkum_value = ws.cell(row=curr_row, column=soll_start_col + 2).value

            subject_name_value = ws.cell(row=curr_row, column=subject_col).value

            teacher_key_value = ws.cell(row=curr_row, column=teacher_key_col).value

            curr_row += 1
            curr_nr_cell = ws.cell(row=curr_row, column=lfd_nr_col)

            if teacher_subject_pair_value is None:
                print(f"warning: teacher subject pair is None for lfd nr {lfd_nr_value}, ignoring")
                continue

            if soll_value is None:
                print(f"warning: soll is None for lfd nr {lfd_nr_value}, ignoring")
                continue

            if teacher_name_value is None:
                print(f"warning: teacher name is None for lfd nr {lfd_nr_value}, ignoring")
                continue

            if soll_value is None:
                print(f"warning: soll is None for lfd nr {lfd_nr_value}, ignoring")
                continue

            if ist_akkum_value is None:
                print(f"warning: ist akkum is None for lfd nr {lfd_nr_value}, ignoring")
                continue

            if subject_name_value is None:
                print(f"warning: subject name is None for lfd nr {lfd_nr_value}, ignoring")
                continue

            soll_info = {
                "lfd_nr": lfd_nr_value,
                "teacher_name": teacher_name_value,
                "teacher_subject_pair": teacher_subject_pair_value,
                "soll": soll_value,
                "ist": ist_akkum_value,
                # "teacher_key": None, # set later
                "teacher_key": teacher_key_value,
                "subject_name": subject_name_value.strip()
            }
            soll_data_array.append(soll_info)

        workbook.close()
        return {
            "class_key": class_key,
            "soll_data_array": soll_data_array,
        }

    def add_soll_data_to_class_subject_teachers(self, all_classes, all_teachers_list):

        for class_obj in all_classes:
            class_key = class_obj["key"]

            if class_key not in self.all_soll_data_dict:
                print(f"warning: class key '{class_key}' not found in stundenerfassung -> skipping")
                continue

            soll_data_array = self.all_soll_data_dict[class_key]

            # TODO proper matching...
            soll_data_to_remove = []

            for soll_data_obj in soll_data_array:
                teacher_name = soll_data_obj['teacher_name']
                soll_teacher_key = soll_data_obj['teacher_key']
                last_part = teacher_name.split(" ")[-1]
                lfd_nr = soll_data_obj['lfd_nr']
                soll_subject_name = soll_data_obj['subject_name']

                teacher_found = False
                for teacher_obj in all_teachers_list:
                    teacher_key = teacher_obj['key']
                    if teacher_key == soll_data_obj['teacher_key']:
                    # last_last_name = teacher_obj['last_name']
                    # if last_last_name == last_part:
                    #     soll_data_obj['teacher_key'] = teacher_obj['key']
                        teacher_found = True
                        print(teacher_obj['key'])
                        break
                if not teacher_found:
                    print(
                        f"warning: teacher '{teacher_name}' from lfd. Nr. {lfd_nr} not found in teacher list -> skipping")
                    soll_data_to_remove.append(soll_data_obj)
                    continue

                # now check subjects
                subject_found = False
                for subject_obj in class_obj['subjects']:
                    class_subject_name = subject_obj['name'].strip()
                    if soll_subject_name == class_subject_name:
                        subject_found = True
                        # add the ist and soll values to the teachers of this subject
                        for teacher_with_hours in subject_obj['teachers_with_hours']:
                            if teacher_with_hours['teacher_key'] == soll_teacher_key:
                                teacher_with_hours['ist'] = soll_data_obj['ist']
                                teacher_with_hours['soll'] = soll_data_obj['soll']
                        break

                if not subject_found:
                    print(f"warning: subject '{soll_subject_name}' from lfd. Nr. {lfd_nr} not found in class '{class_key}' -> skipping")
                    soll_data_to_remove.append(soll_data_obj)
                    continue

            for soll_data_obj in soll_data_to_remove:
                soll_data_array.remove(soll_data_obj)

        pass


if __name__ == '__main__':
    excel_extractor = ExcelExtractorStundenerfassung(
        ['example/Std-erfassung_Erz24A_1. und 2. Sj.xlsx'],
        0
    )
    excel_extractor.read_all()
