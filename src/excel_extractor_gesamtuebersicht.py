# https://openpyxl.readthedocs.io/en/stable/tutorial.html
import os

import openpyxl

from src.logger import Logger

MAPPINGS_SHEET_COLORS = "farben"
MAPPINGS_SHEET_CLASSES = "klassen"
MAPPINGS_SHEET_SUBJECTS = "fächer"

SHEET_OVERVIEW_PLAN = 'Plan'

# true: we only use every 2nd slot
# false: use every slow
ONLY_USE_BLOCKS_OF_TWO = True
# https://openpyxl.readthedocs.io/en/3.1.3/_modules/openpyxl/styles/colors.html#Color
KNOWN_COLOR_BLACK = '00000000'
KNOWN_COLOR_WHITE = '00FFFFFF'


# teachers start at row 10, col B - E
# subjects start at col J, row 6
# class is in col J, row 3
#  subjects from other classes are separated by empty cells or start with (...)
# required lessons in UE for subjects are in row 9 (below subjects)

# dieser helper liest die datei  SJ 25-26_Gesamtübersicht Einsatz Lehrkräfte EA Halle 2025-05-21_3 aus
# darin sind alle lehrer zu finden, alle klassen, fächer und wv stunden ein lehrere für ein fach in einer klasse unterrichtet.
class ExcelExtractorGesamtuebersicht:
    def __init__(self, excel_file_all_classes_path, excel_file_mappings, excel_file_plan_availability_path):
        self.log_name = "ExcelExtractorGesamtuebersicht"
        self.excel_file_all_classes_path = excel_file_all_classes_path
        self.excel_file_mappings = excel_file_mappings
        self.excel_file_plan_availability_path = excel_file_plan_availability_path

        self.color_legend_teacher_availability = None
        self.all_teachers_list = None
        self.all_teachers_dict = None
        self.all_classes = None
        self.subject_name_to_key_dict = None

    def read_all(self):
        Logger.debug(f"[{self.log_name}] reading excel file {self.excel_file_all_classes_path}")
        wb_overview_classes = openpyxl.load_workbook(self.excel_file_all_classes_path, read_only=False)
        worksheet_blacklist = ["variablen", "kumuliert"]
        all_relevant_work_sheets = []

        Logger.debug(f"[{self.log_name}] scanning for relevant worksheets, blacklisted: {worksheet_blacklist} ...")
        for sheet in wb_overview_classes:
            worksheet_name = sheet.title.lower()
            is_relevant = True
            for blacklisted_name in worksheet_blacklist:
                if blacklisted_name in worksheet_name:
                    is_relevant = False
                    break

            if not is_relevant:
                continue
            all_relevant_work_sheets.append(sheet)
            Logger.debug(
                f"[{self.log_name}] found relevant sheet: {sheet.title} (index: {all_relevant_work_sheets.index(sheet)})")

        self.all_teachers_list = []
        self.all_teachers_dict = {}
        self.all_classes = []
        self.subject_name_to_key_dict = {}
        first_pass = True

        # ws = wb['KP 25_26']
        for ws in all_relevant_work_sheets:
            Logger.debug(f"[{self.log_name}] extracting all data from sheet {ws.title}...")
            data = self.extract_all_data_from_sheet(ws)
            for teacher_obj in data["all_teachers_list"]:
                teacher_key = teacher_obj["key"]
                if teacher_key not in self.all_teachers_dict:
                    if not first_pass:
                        Logger.error(
                            f"[{self.log_name}] [{ws.title}] teacher '{teacher_key}' was not in the first table! Check if this is a mistake!")

                    self.all_teachers_dict[teacher_key] = teacher_obj
                    self.all_teachers_list.append(teacher_obj)
                else:
                    pass

            first_pass = False

            for class_obj in data["all_classes"]:
                self.all_classes.append(class_obj)
                for subject_obj in class_obj["subjects"]:
                    self.subject_name_to_key_dict[subject_obj["name"]] = None

            # break
        print(f"found {len(self.all_teachers_list)} teachers")
        print(f"found {len(self.all_classes)} classes")
        #
        # # write to json file
        # teacher_data_json = json.dumps(all_teachers_list, indent=4)
        # with open('example/all_teachers.json', 'w') as outfile:
        #     outfile.write(teacher_data_json)

        # read from json file
        # with open('example/all_teachers.json', 'r') as infile:
        #     all_teachers_list = json.load(infile)
        #
        #     for teacher_obj in all_teachers_list:
        #         teacher_key = teacher_obj["key"]
        #         all_teachers_dict[teacher_key] = teacher_obj

        # check if mappings file exists
        self._ensure_mappings_file_exists_and_required_sheets(self.excel_file_mappings, self.all_classes,
                                                              self.subject_name_to_key_dict)

        wb_mappings = openpyxl.load_workbook(self.excel_file_mappings, read_only=False, data_only=True)

        self.color_legend_teacher_availability = self.extract_class_and_teacher_available_color_mapping(wb_mappings)

        class_key_to_full_name_dict = self.extract_class_mapping_from_sheet(wb_mappings, self.all_classes)
        self.add_class_key_to_all_classes(self.all_classes, class_key_to_full_name_dict)

        self.extract_subject_key_to_subject_mapping_from_sheet(wb_mappings, self.subject_name_to_key_dict)

        wb_plan_preferences = openpyxl.load_workbook(self.excel_file_plan_availability_path, read_only=False,
                                                     data_only=True)

        # this is only for reading, no writing!
        self.extract_teachers_availability_and_prefs_from_sheet(wb_plan_preferences,
                                                                self.all_teachers_dict,
                                                                self.all_teachers_list,
                                                                self.color_legend_teacher_availability,
                                                                self.all_classes)

        self._teacher_sanity_checks(self.all_teachers_dict, self.all_teachers_list, self.all_classes)

        # sets which slots should be filled and what should be ignored
        self.make_teacher_availability_and_prefs_canonical(self.all_teachers_list,
                                                           self.color_legend_teacher_availability)

        all_table_data_dict = self.extract_current_plan_from_sheet(wb_plan_preferences,
                                                                   self.color_legend_teacher_availability)


        # this also removes empty classes (no subjects)
        self._validate_class_and_teachers(self.all_classes, self.all_teachers_list, self.all_teachers_dict)

        self._set_class_plans(all_table_data_dict, self.all_classes)

        wb_overview_classes.close()
        wb_plan_preferences.close()
        # we are only allowed to write to "PLAN" sheet
        Logger.log(f"[{self.log_name}] finished reading excel files")

    def _ensure_mappings_file_exists_and_required_sheets(self, excel_file_mappings, all_classes,
                                                         subject_name_to_key_dict):

        if not os.path.exists(excel_file_mappings):
            # then create it
            Logger.log(f"[{self.log_name}] creating mappings file '{excel_file_mappings}'")
            wb_mappings = openpyxl.Workbook()
        else:
            Logger.log(
                f"[{self.log_name}] mappings file '{excel_file_mappings}' already exists, checking required sheets")
            wb_mappings = openpyxl.load_workbook(excel_file_mappings, read_only=False, data_only=True)

        some_changed = False

        if MAPPINGS_SHEET_COLORS not in wb_mappings.sheetnames:
            Logger.log(f"[{self.log_name}] creating color mapping sheet ('{MAPPINGS_SHEET_COLORS}') in mappings file")

            # create color mapping
            ws_color_mapping = wb_mappings.create_sheet(MAPPINGS_SHEET_COLORS)

            ws_color_mapping.cell(row=2,
                                  column=1).value = '10 (verschiedene) Farben sind für "kann nur da hin" vorgesehen, 10 für "kann da nicht hin"'
            ws_color_mapping.cell(row=9, column=1).value = 'Farblegende Programm (genau die Farben!)'
            ws_color_mapping.cell(row=9, column=4).value = 'Erklärung'  # D5
            ws_color_mapping.cell(row=9, column=5).value = 'Hinweis'  # E5

            default_allowed_color = "92D050"
            default_not_allowed_color = "FF0000"
            # set bg color to green
            # excel uses #aarrggbb for colors, use fgColor in fill!! (bg is for patterns)
            ws_color_mapping.cell(row=10, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=11, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=12, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=13, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=14, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=15, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=16, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=17, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=18, column=1).fill.fgColor.rgb = default_allowed_color
            ws_color_mapping.cell(row=19, column=1).fill.fgColor.rgb = default_allowed_color

            # not allowed
            ws_color_mapping.cell(row=20, column=1).fill.fgColor.rgb = default_not_allowed_color
            ws_color_mapping.cell(row=21, column=1).fill.fgColor.rgb = "C91BBD"
            ws_color_mapping.cell(row=22, column=1).fill.fgColor.rgb = "F7C7AC"
            ws_color_mapping.cell(row=23, column=1).fill.fgColor.rgb = "FF5050"
            ws_color_mapping.cell(row=24, column=1).fill.fgColor.rgb = "C00000"
            ws_color_mapping.cell(row=25, column=1).fill.fgColor.rgb = "FF6600"
            ws_color_mapping.cell(row=26, column=1).fill.fgColor.rgb = default_not_allowed_color
            ws_color_mapping.cell(row=27, column=1).fill.fgColor.rgb = default_not_allowed_color
            ws_color_mapping.cell(row=28, column=1).fill.fgColor.rgb = default_not_allowed_color
            ws_color_mapping.cell(row=29, column=1).fill.fgColor.rgb = default_not_allowed_color

            some_changed = True
        else:
            Logger.log(
                f"[{self.log_name}] color mapping sheet ('{MAPPINGS_SHEET_COLORS}') already exists in mappings file")

        if MAPPINGS_SHEET_CLASSES not in wb_mappings.sheetnames:
            Logger.log(f"[{self.log_name}] creating color mapping sheet ('{MAPPINGS_SHEET_CLASSES}') in mappings file")

            # create class mapping (key to full name)
            ws_class_mapping = wb_mappings.create_sheet(MAPPINGS_SHEET_CLASSES)

            ws_class_mapping.cell(row=1,
                                  column=1).value = 'Die Klassennamen Teile sind in der Gesamtübersicht zu finden (3 Zeile)'
            ws_class_mapping.cell(row=3,
                                  column=1).value = 'Kürzel frei lassen, damit nicht beachtet wird (Plan wird nicht ausgefüllt)'

            ws_class_mapping.cell(row=9, column=1).value = 'Klassen Kürzel'
            ws_class_mapping.cell(row=9, column=2).value = 'Klassenname 1'
            ws_class_mapping.cell(row=9, column=3).value = 'Klassenname 2'
            ws_class_mapping.cell(row=9, column=4).value = 'Klassenname 3'

            curr_row = 10

            for class_obj in all_classes:
                class_key = class_obj['key']
                class_name_parts = class_obj['name_fields']
                for i, part in enumerate(class_name_parts):
                    ws_class_mapping.cell(row=curr_row, column=1 + i).value = part

                curr_row += 1

            some_changed = True

        else:
            Logger.log(
                f"[{self.log_name}] class mapping sheet ('{MAPPINGS_SHEET_CLASSES}') already exists in mappings file")

        if MAPPINGS_SHEET_SUBJECTS not in wb_mappings.sheetnames:
            Logger.log(f"[{self.log_name}] creating color mapping sheet ('{MAPPINGS_SHEET_SUBJECTS}') in mappings file")

            # create subject key to subject mapping
            ws_subjects_mapping = wb_mappings.create_sheet("fächer")

            ws_subjects_mapping.cell(row=1, column=1).value = 'Der Fach Name ist der volle Name aus der Gesamtübersicht'

            ws_subjects_mapping.cell(row=9, column=1).value = 'Fach Kürzel'
            ws_subjects_mapping.cell(row=9, column=2).value = 'Fach Name'

            subject_name_to_key_dict_items = list(subject_name_to_key_dict.items())

            # subject_name_to_key_dict_items.sort(key=lambda x: x[1])

            for i, subject_name_to_key_dict_item in enumerate(subject_name_to_key_dict_items):
                subject_name = subject_name_to_key_dict_item[0]
                ws_subjects_mapping.cell(row=10 + i, column=2).value = subject_name

            some_changed = True

        else:
            Logger.log(
                f"[{self.log_name}] color mapping sheet ('{MAPPINGS_SHEET_SUBJECTS}') already exists in mappings file")

        if some_changed:
            wb_mappings.save(excel_file_mappings)

    def extract_subject_key_to_subject_mapping_from_sheet(self, wb_mappings, subject_name_to_key_dict):
        mapping_worksheet = wb_mappings[MAPPINGS_SHEET_SUBJECTS]

        # all_known_subjects = set()
        #
        # for class_obj in all_classes:
        #     subjects_list = class_obj['subjects']
        #     for subject_obj in subjects_list:
        #         subject_name = subject_obj['name']
        #         all_known_subjects.add(subject_name)

        # subject short | subject name
        start_col = 1
        start_row = 10
        # max_row = 1000  # until we find empty row, but to be safe here
        max_row = mapping_worksheet.max_row + 1
        error_count = 0
        found_subject_names = []

        for row_j in range(start_row, max_row):
            subject_short_cell = mapping_worksheet.cell(row=row_j, column=start_col)
            subject_name_cell = mapping_worksheet.cell(row=row_j, column=start_col + 1)

            if subject_short_cell.value is None:
                raise Exception(
                    f"[{self.log_name}][mapping excel file] subject short form is None in row {row_j}, column {start_col} (name: {subject_name_cell.value})")

            subject_short = subject_short_cell.value
            subject_name = subject_name_cell.value

            # subject_name_to_key_dict contains all subject names with None as key
            if subject_name not in subject_name_to_key_dict:
                Logger.warn(
                    f"[{self.log_name}][mapping excel file] no class has subject '{subject_name}' in sheet '{MAPPINGS_SHEET_SUBJECTS}'")
                continue

            found_subject_names.append(subject_name)

            if subject_name_to_key_dict[subject_name] is not None:
                Logger.warn(
                    f"[{self.log_name}][mapping excel file] subject '{subject_name}' already has a short form in sheet '{MAPPINGS_SHEET_SUBJECTS}'")
                error_count += 1

            subject_name_to_key_dict[subject_name] = subject_short

        if error_count > 0:
            raise Exception(
                f"[{self.log_name}][mapping excel file] {error_count} errors in subject mapping, sheet '{MAPPINGS_SHEET_SUBJECTS}'")

        # not check if all know subjects have a short form
        new_lines_count = 0
        known_subjects = list(subject_name_to_key_dict.keys())
        for subject_name in known_subjects:
            subject_short_form = subject_name_to_key_dict[subject_name]

            if subject_short_form is None:
                Logger.error(
                    f"[{self.log_name}][mapping excel file] subject '{subject_name}' has no short form in sheet '{MAPPINGS_SHEET_SUBJECTS}', adding it for you!")
                error_count += 1

                mapping_worksheet.cell(row=max_row + new_lines_count, column=start_col).value = None
                mapping_worksheet.cell(row=max_row + new_lines_count, column=start_col + 1).value = subject_name
                new_lines_count += 1

        if new_lines_count > 0:
            wb_mappings.save(self.excel_file_mappings)

        if error_count > 0:
            raise Exception(f"[{self.log_name}][mapping excel file] {error_count} errors in subject mapping")

    def extract_class_and_teacher_available_color_mapping(self, wb_mappings):
        # excel uses aarrggbb for colors, use fgColor in fill!! (bg is for patterns)
        ws_color_legend = wb_mappings[MAPPINGS_SHEET_COLORS]
        mapping_allowed_row_start = 10

        max_num_colors = 10

        # in individual teachers sheets
        allowed_colors = set()
        # this will also be used as not process color
        not_allowed_colors = set()

        for i in range(max_num_colors):
            mapping_allowed_cell = ws_color_legend.cell(row=mapping_allowed_row_start + i - 1, column=1)
            allowed_color = self.get_cell_color(mapping_allowed_cell)
            allowed_colors.add(allowed_color)

        mapping_not_allowed_row_start = 20

        # in individual teachers sheets
        for i in range(max_num_colors):
            mapping_allowed_cell = ws_color_legend.cell(row=mapping_not_allowed_row_start + i - 1, column=1)
            not_allowed_color = self.get_cell_color(mapping_allowed_cell)
            not_allowed_colors.add(not_allowed_color)

        Logger.debug(f"[{MAPPINGS_SHEET_COLORS}] allowed colors: {allowed_colors}")
        Logger.debug(f"[{MAPPINGS_SHEET_COLORS}] not allowed colors: {not_allowed_colors}")

        return {
            "allowed_bg_colors_set": allowed_colors,
            "not_allowed_bg_colors_set": not_allowed_colors
        }

    def extract_class_mapping_from_sheet(self, wb_mappings, all_classes):
        ws_class_mapping = wb_mappings[MAPPINGS_SHEET_CLASSES]
        # mapping from class keys to class full names (3 columns)
        class_key_to_full_name_dict = {}

        # class key | class name part 1 | class name part 2 (optional) | class name part 3 (optional)
        start_col = 1
        start_row = 10
        # max_row = 100  # until we find empty row, but to be safe here
        max_row = ws_class_mapping.max_row

        existing_mapping_but_no_key_set = set()

        for row_j in range(start_row, max_row + 1):
            class_key = ws_class_mapping.cell(row=row_j, column=start_col)
            class_name_part_1 = ws_class_mapping.cell(row=row_j, column=start_col + 1)
            class_name_part_2 = ws_class_mapping.cell(row=row_j, column=start_col + 2)
            class_name_part_3 = ws_class_mapping.cell(row=row_j, column=start_col + 3)

            name_full = ""

            if class_name_part_1.value is not None:
                name_full += class_name_part_1.value

            if class_name_part_2.value is not None:
                name_full += f" {class_name_part_2.value}"

            if class_name_part_3.value is not None:
                name_full += f" {class_name_part_3.value}"

            if class_key.value is None:
                if name_full == "":  # ignore empty lines
                    continue

                existing_mapping_but_no_key_set.add(name_full)
                Logger.warn(
                    f"[{self.log_name}][mappings excel file] class mapping in row {row_j} is empty, IGNORING CLASS (class name parts: {name_full})")
                continue

            if class_name_part_1.value is None:
                raise Exception(
                    f"[{self.log_name}][mappings excel file] class name part 1 is required in worksheet '{MAPPINGS_SHEET_CLASSES} for class key '{class_key.value}'")

            if class_key.value in class_key_to_full_name_dict:
                raise Exception(
                    f"[{self.log_name}][mappings excel file] class key '{class_key.value}' is already in class mapping")

            class_key_to_full_name_dict[class_key.value] = {
                "name_part_1": class_name_part_1.value,
                "name_part_2": class_name_part_2.value,
                "name_part_3": class_name_part_3.value,
                "name_full": name_full
            }

        Logger.debug(
            f"[{self.log_name}][mappings excel file] class mapping ({len(class_key_to_full_name_dict)}): {class_key_to_full_name_dict}")

        # now check if we have an entry for each class

        not_found_classes = []

        for class_obj in all_classes:
            class_name_single_line = class_obj['name_single_line']
            found_class_key = False

            for class_key, class_name_infos in class_key_to_full_name_dict.items():
                if class_name_infos['name_full'] == class_name_single_line:
                    found_class_key = True
                    break
            if not found_class_key:

                if class_name_single_line in existing_mapping_but_no_key_set:
                    continue

                not_found_classes.append(class_obj)
                Logger.error(
                    f"[{self.log_name}][mappings excel file] class '{class_name_single_line}' not found in class mapping, will be inserted for you!")

        if len(not_found_classes) > 0:

            for i, class_obj in enumerate(not_found_classes):
                class_name_parts = class_obj['name_fields']

                ws_class_mapping.cell(row=max_row + 1 + i, column=1).value = None  # must be filled by the user
                for j, part in enumerate(class_name_parts):
                    ws_class_mapping.cell(row=max_row + 1 + i, column=2 + j).value = part

            wb_mappings.save(self.excel_file_mappings)
            raise Exception(
                f"[{self.log_name}][mappings excel file] {len(not_found_classes)} classes not found in class mapping.")

        return class_key_to_full_name_dict

    def get_max_slots_per_day(self):
        first_teacher = self.all_teachers_list[0]
        max_slots_per_day = len(first_teacher['availability_preference_table'][0])
        return max_slots_per_day

    def get_max_days(self):
        return len(self.all_teachers_list[0]['availability_preference_table'])

    def write_timetable_solution_to_excel_impl(self, output_file_path,
                                               timetable_solution_tuples_for_classes_with_at_least_one_lesson):

        source_file = self.excel_file_plan_availability_path

        # if the output_file_path ends with .xlsm, force it to be .xlsx
        if output_file_path.endswith(".xlsm"):
            output_file_path = output_file_path[:-5] + ".xlsx"

        # create copy of file
        # shutil.copyfile(source_file, output_file_path)

        wb_plan_output = openpyxl.load_workbook(source_file, read_only=False)

        ws_plan_out = wb_plan_output['Plan']

        for timetable_solutions_tuple in timetable_solution_tuples_for_classes_with_at_least_one_lesson:
            class_obj = timetable_solutions_tuple[0]
            print(f"writing timetable solution for class '{class_obj['key']}' ('{class_obj['name_single_line']}')")
            timetable_solution_dict = timetable_solutions_tuple[1]
            table_start_coord_tuple = timetable_solutions_tuple[2]

            # top left cell of class timetable
            class_key_row_in_plan = table_start_coord_tuple['start_row']
            class_key_col_in_plan = table_start_coord_tuple['start_col']

            num_days = len(timetable_solution_dict)

            for day_index in range(num_days):
                timetable_solution_slot_list = timetable_solution_dict[day_index]
                for slot_index, timetable_solution_entry in enumerate(timetable_solution_slot_list):
                    if timetable_solution_entry is None:
                        continue
                    # row + 1 are the dates
                    plan_cell = ws_plan_out.cell(column=class_key_col_in_plan + 1 + day_index,
                                                 row=class_key_row_in_plan + 2 + slot_index)
                    teacher_key = timetable_solution_entry['teacher_key']
                    subject_key = timetable_solution_entry['subject_key']
                    class_key = timetable_solution_entry['class_key']
                    solution_value = None

                    if subject_key in self.subject_name_to_key_dict:
                        solution_value = f"{self.subject_name_to_key_dict[subject_key]}({teacher_key})"
                    else:
                        solution_value = f"{subject_key}({teacher_key})"
                        print(
                            f"warning: subject key '{subject_key}' not found in subject name to key dict -> using as is")

                    plan_cell.value = solution_value

        wb_plan_output.save(output_file_path)
        # wb_plan_output.close()

    def get_excel_rect_data_as_array(self, ws, min_row, max_row, min_col, max_col):
        array = []
        for row in range(min_row, max_row + 1):
            array.append([])
            all_none = True
            for col in range(min_col, max_col + 1):
                data = ws.cell(row=row, column=col).value
                if data is not None:
                    all_none = False
                array[-1].append(data)

            if all_none:
                # remove last row
                del array[-1]
                break

        return array

    def get_cell_range_from_merged_cells(self, ws, cell):
        merged_cells = ws.merged_cells.ranges
        for merged_cell in merged_cells:
            if cell.coordinate in merged_cell:
                return merged_cell.bounds  # col, row, col, row

    def extract_classes_with_data_from_sheet(self, ws):
        start_row = 3
        start_col = 10  # J
        subjects_start_row = 6
        subjects_hours_year = 7
        subjects_hours_term = 9  # for one half year (one term/semester)

        all_classes = []
        curr_class = ws.cell(row=start_row, column=start_col)  # J3

        # some classes have 1 / 2 / 3 rows ...
        # the subjects for all classes start in row 6
        while curr_class.value is not None:
            #       col, row, col, row
            # e.g. (10,  3,   32,  3)
            range_class_name = self.get_cell_range_from_merged_cells(ws, curr_class)
            Logger.debug(
                f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' range: col ({range_class_name[0]} - {range_class_name[2]}), row ({range_class_name[1]} - {range_class_name[3]}))")
            class_name_lines = []

            class_name_lines.append(curr_class.value.strip())
            # extract class data
            if range_class_name[1] == range_class_name[3]:
                # whole class in one merged cell (same row)
                # so check the next rows until we get to row 6 (row 6 are the subjects)
                class_data_2 = ws.cell(row=range_class_name[1] + 1, column=range_class_name[0])
                range_class_data_2 = self.get_cell_range_from_merged_cells(ws, class_data_2)
                class_name_lines.append(class_data_2.value.strip() if class_data_2.value is not None else "")
                Logger.debug(f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' name2: {class_data_2.value}")

                if range_class_data_2[1] == range_class_data_2[3]:
                    # one row
                    class_data_3 = ws.cell(row=range_class_data_2[1] + 1, column=range_class_data_2[0])
                    range_class_data_3 = self.get_cell_range_from_merged_cells(ws, class_data_3)
                    if class_data_3.value is None:
                        class_name_lines.append("")
                        Logger.debug(
                            f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' name2: {class_data_2.value}, name3: None")
                    else:
                        class_name_lines.append(class_data_3.value.strip())
                        Logger.debug(
                            f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' name2: {class_data_2.value}, name3: {class_data_3.value}")

                else:
                    # data is a merged cell with more than one row -> there is no data 3
                    pass
            else:
                # data 1 is more than one row high
                if range_class_name[3] >= subjects_start_row:
                    # we only have one data in a big merged cell
                    raise Exception("TODO")
                else:
                    # data 1 is 2 rows high, so we can have data 3
                    class_data_3 = ws.cell(row=range_class_name[1] + 1, column=range_class_name[0])
                    range_class_data_3 = self.get_cell_range_from_merged_cells(ws, class_data_3)
                    if class_data_3.value is None:
                        class_name_lines.append("")
                        Logger.debug(
                            f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' (name1 is >= 1 row heigh), name3: None")
                    else:
                        class_name_lines.append(class_data_3.value.strip())
                        Logger.debug(
                            f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' (name1 is >= 1 row heigh), name3: {class_data_3.value}")

            class_obj = {
                "name": str.join("\n", class_name_lines).strip(),
                "name_fields": class_name_lines,
                "name_single_line": str.join(" ", class_name_lines).strip(),
                "col_range": [range_class_name[0], range_class_name[2]],
                "subjects": [],  # {"name", "col", "coord"}
                "fake_subjects": []
            }

            # TODO infinite loop...
            # print("TODO infinite loop...")
            # _first_cell = ws.cell(row=range_class_name[0], column=range_class_name[1])
            # _last_cell = ws.cell(row=range_class_name[2], column=range_class_name[1])
            #
            # print(f"info: {class_obj['name_single_line']} has column range: {class_obj['col_range']} [{_first_cell.coordinate} - {_last_cell.coordinate}]")

            all_classes.append(class_obj)

            if class_obj['name_single_line'] == 'Erzieher ':
                Logger.warn(
                    f"[{self.log_name}] [{ws.title}] class '{curr_class.value}' special case because merged cell is wrong ...")
                # print("TODO 'Erzieher ' ...")
                curr_class = ws.cell(row=start_row, column=range_class_name[2] + 2)
            else:
                curr_class = ws.cell(row=start_row, column=range_class_name[2] + 1)

        # now we have all classes
        # extract subjects for each class
        for class_obj in all_classes:
            class_col_range = class_obj["col_range"]
            for col in range(class_col_range[0], class_col_range[1] + 1):
                subject_cell = ws.cell(row=subjects_start_row, column=col)

                # some misc cell, e.g. sum of hours or something
                if subject_cell.value is None:
                    continue

                # we have fake subjects e.g. managing a class
                subject_hours_total = ws.cell(row=subjects_hours_year, column=col)
                subject_hours_term = ws.cell(row=subjects_hours_term, column=col)

                subject_obj = {
                    "name": subject_cell.value,
                    "col": subject_cell.column,
                    "coord": subject_cell.coordinate,
                    "hours_total": subject_hours_total.value,
                    "hours_term": subject_hours_term.value,
                    "teachers_with_hours": []  # {"teacher_key", "hours"}
                }

                if subject_hours_term.value is None:
                    # not a real subject...
                    class_obj["fake_subjects"].append(subject_obj)
                else:
                    class_obj["subjects"].append(subject_obj)
                    Logger.debug(
                        f"[{self.log_name}] [{ws.title}] class '{class_obj['name_single_line']}' found subject: '{subject_cell.value}'")

        return all_classes

    def extract_and_set_teacher_hours_from_sheet(self, ws, all_classes, all_teachers_list, teacher_row_start):
        for class_obj in all_classes:
            subject_objs = class_obj["subjects"]
            for subject_obj in subject_objs:
                subject_col = subject_obj["col"]

                for row_index, teacher_obj in enumerate(all_teachers_list):
                    hours_for_teacher_cell = ws.cell(row=teacher_row_start + row_index, column=subject_col)

                    if hours_for_teacher_cell.value is None:
                        continue

                    subject_obj["teachers_with_hours"].append({
                        "teacher_key": teacher_obj["key"],
                        "teacher_full_name": teacher_obj['teacher_full_name'],
                        "hours": hours_for_teacher_cell.value
                    })
                    Logger.debug(
                        f"[{self.log_name}] [{ws.title}] [{class_obj['name_single_line']}] teacher '{Logger.get_teacher_full(teacher_obj)}' found '{hours_for_teacher_cell.value}' hours for subject '{subject_obj['name']}' [cell: {Logger.get_cell_full_coord(hours_for_teacher_cell)}]")

    def get_all_teachers_from_rect_data(self, teacher_datas):
        all_teachers_dict = {}
        all_teachers_list = []

        for teacher_data in teacher_datas:
            i = 0
            contract_form = teacher_data[i]
            i += 1
            last_name = teacher_data[i]
            i += 1
            first_name = teacher_data[i]
            i += 1
            teacher_key = teacher_data[i]
            teacher_obj = {
                "contract_form": contract_form,
                "last_name": last_name,
                "first_name": first_name,
                "teacher_full_name": f"{first_name} {last_name}",
                "key": teacher_key,
                "availability_preference_table": None
            }
            all_teachers_dict[teacher_key] = teacher_obj
            all_teachers_list.append(teacher_obj)

        return all_teachers_list, all_teachers_dict

    def extract_all_data_from_sheet(self, ws):
        # key is "teacher key" (or nummer in excel/german)
        # value is hash with data
        teachers = dict()

        teacher_row_start = 10
        # teacher_end_row = 100
        teacher_end_row = ws.max_row
        teacher_start_col = 2
        teacher_end_col = 5
        Logger.debug(f"[{self.log_name}] [{ws.title}] getting data in rect: min_row, max_row, min_col, max_col: "
                     f"{teacher_row_start}, {teacher_end_row}, {teacher_start_col}, {teacher_end_col}")
        teacher_datas = self.get_excel_rect_data_as_array(ws, teacher_row_start, teacher_end_row, teacher_start_col,
                                                          teacher_end_col)

        Logger.debug(f"[{self.log_name}] [{ws.title}] converting rect teacher data to teacher objs")
        all_teachers_list, all_teachers_dict = self.get_all_teachers_from_rect_data(teacher_datas)
        Logger.debug(f"[{self.log_name}] [{ws.title}] found {len(all_teachers_list)} teachers")

        Logger.debug(f"[{self.log_name}] [{ws.title}] extracting classes with data from sheet")
        all_classes = self.extract_classes_with_data_from_sheet(ws)
        Logger.debug(
            f"[{self.log_name}] [{ws.title}] extracted {len(all_classes)} classes: {list(map(lambda x: x['name_single_line'], all_classes))}")

        Logger.debug(f"[{self.log_name}] [{ws.title}] setting hours for teachers...")
        self.extract_and_set_teacher_hours_from_sheet(ws, all_classes, all_teachers_list, teacher_row_start)

        return {
            "all_teachers_list": all_teachers_list,
            "all_teachers_dict": all_teachers_dict,
            "all_classes": all_classes,
        }

    def get_cell_color(self, cell):
        # excel uses #aarrggbb for colors, use fgColor in fill!! (bg is for patterns)
        return cell.fill.fgColor

    def extract_and_set_single_teacher_availability_preferences_from_sheet(self, ws, teacher_obj,
                                                                           color_legend_teacher_availability):
        start_col = 3  # C, here is also the name of the class and time slots
        end_col = 7  # G

        start_row = 2  # first time slot
        end_row = 10  # last time slot

        # true: only read every 2nd entry because we only process blocks of 2 entries
        # false: read every entry
        # only_use_blocks = ONLY_USE_BLOCKS_OF_TWO

        date_cells = []

        # C2 - G2 are the dates
        for i in range(start_col, end_col + 1):
            date_cell = ws.cell(row=start_row, column=i)
            date_cells.append(date_cell.value)

        preferences_cells = []  # 2d, column wise

        has_known_color = None

        # real data
        for col_i in range(start_col, end_col + 1):

            day_slots = []
            for row_j in range(start_row + 1, end_row + 1, 2 if ONLY_USE_BLOCKS_OF_TWO else 1):

                pref_cell = ws.cell(row=row_j, column=col_i)
                # print(f"{teacher_obj['key']}, {col_i}, {row_j}")

                cell_obj = {
                    "slot_index": pref_cell.row - (start_row + 1),
                    "class_key": pref_cell.value,  # can be None if empty
                    "color": self.get_cell_color(pref_cell),
                    "allowed": None  # will be set later (black/white list)
                }

                if (cell_obj["color"] in color_legend_teacher_availability['not_allowed_bg_colors_set'] or
                        cell_obj["color"] in color_legend_teacher_availability['allowed_bg_colors_set']):
                    if has_known_color is not None:
                        if cell_obj["color"] != has_known_color:
                            # we only allow one color... white or blacklist

                            # see https://openpyxl.readthedocs.io/en/3.1.3/_modules/openpyxl/styles/colors.html#Color
                            if cell_obj["color"].rgb == KNOWN_COLOR_BLACK or cell_obj["color"].rgb != KNOWN_COLOR_WHITE:
                                pass  # this is ok
                            else:
                                Logger.error(
                                    f"[{self.log_name}][Sheet {teacher_obj['key']}] teacher {Logger.get_teacher_full(teacher_obj)} has both allowed and not allowed colors")
                                return False

                    if cell_obj["color"] in color_legend_teacher_availability["not_allowed_bg_colors_set"]:
                        has_known_color = cell_obj["color"]
                    if cell_obj["color"] in color_legend_teacher_availability["allowed_bg_colors_set"]:
                        has_known_color = cell_obj["color"]

                day_slots.append(cell_obj)
            preferences_cells.append(day_slots)
            # print(f"{teacher_obj['key']}")

        teacher_obj['availability_preference_table'] = preferences_cells
        # print(f"{teacher_obj['key']}")

        return True

    # obsolete
    def __extract_subject_mapping_from_sheet(self, ws_dozenten, all_teachers_dict):
        start_col = 1
        max_col = 100
        # no end col, we read until we find a class without a trailing dot
        start_row = 1
        max_row = 100

        subject_teacher_pair_by_class_name_dict = {}

        for col_i in range(start_col, max_col):
            class_name_cell = ws_dozenten.cell(row=start_row, column=col_i)
            class_name = class_name_cell.value

            subject_teacher_pair = []
            subject_teacher_pair_by_class_name_dict[class_name] = subject_teacher_pair

            for row_j in range(start_row + 1, max_row):
                subject_teacher_pair_cell = ws_dozenten.cell(row=row_j, column=col_i)
                value = subject_teacher_pair_cell.value

                # value is a pair, e.g. Eth(Bor) -> Eth, Bor
                if value is not None:
                    # print(f"{subject_teacher_pair_cell.coordinate} = {value}")

                    # if the value does not contain "(", ignore
                    if "(" not in value:
                        print(f"warning: subject teacher pair does not contain '(' for class {class_name}")
                        continue

                    subject, teacher = value.split("(")
                    teacher_key = teacher.strip(")")

                    if teacher_key == "":
                        print(f"warning: teacher key is empty for subject {subject} for class {class_name}")
                        continue
                    if teacher_key not in all_teachers_dict:
                        print(
                            f"warning: teacher key {teacher_key} not found for subject {subject} for class {class_name}")
                        continue

                    subject_teacher_pair.append({"subject": subject, "teacher_key": teacher_key})

        return subject_teacher_pair_by_class_name_dict

    def _remove_teacher_and_from_class_subjects(self, teacher_obj, all_teachers_dict, all_teachers_list, all_classes):

        key = teacher_obj["key"]
        all_teachers_list.remove(teacher_obj)
        del all_teachers_dict[key]

        # check if the teacher is associated with a class
        for class_obj in all_classes:
            subject_objs = class_obj["subjects"]
            for subject_obj in subject_objs:
                teacher_with_hours_to_remove = []
                for teacher_with_hours in subject_obj["teachers_with_hours"]:
                    if teacher_with_hours["teacher_key"] == key:
                        teacher_with_hours_to_remove.append(teacher_with_hours)

                for teacher_with_hours in teacher_with_hours_to_remove:
                    subject_obj["teachers_with_hours"].remove(teacher_with_hours)
                    Logger.log(
                        f"[{self.log_name}] teacher '{Logger.get_teacher_full(teacher_obj)}' is associated with class '{class_obj['key']}' with subject '{subject_obj['name']}' -> teacher was removed from class with subject")

    def extract_teachers_availability_and_prefs_from_sheet(self, wb_prefs, all_teachers_dict, all_teachers_list,
                                                           color_legend_teacher_availability, all_classes):
        # dozenten_worksheet = wb_prefs["Dozenten"]

        # not needed anymore, we now use extract_subject_key_to_subject_mapping_from_sheet
        # mapping = extract_subject_mapping_from_sheet(dozenten_worksheet, all_teachers_dict)

        # each teacher has a separate sheet with it's key
        # tuple of worksheet and teacher key
        relevant_work_sheets = []

        all_found_worksheet_teacher_names = set()

        for sheet in wb_prefs:
            worksheet_name = sheet.title  # to lower?

            if worksheet_name in all_teachers_dict:
                relevant_work_sheets.append((sheet, worksheet_name))
                all_found_worksheet_teacher_names.add(worksheet_name)  # this is the teacher name

        teachers_to_remove = []
        for key, teacher_obj in all_teachers_dict.items():
            if teacher_obj["key"] not in all_found_worksheet_teacher_names:
                Logger.warn(
                    f"[{self.log_name}] teacher '{Logger.get_teacher_full(teacher_obj)}' (contract form: {teacher_obj['contract_form']}) has no worksheet/plan -> WILL BE IGNORED (and all connections to classes with subjects will be removed!)")
                teachers_to_remove.append(teacher_obj)

        for teacher_obj in teachers_to_remove:
            self._remove_teacher_and_from_class_subjects(teacher_obj, all_teachers_dict, all_teachers_list, all_classes)

        error_count = 0

        for sheet, teacher_key in relevant_work_sheets:
            teacher_obj = all_teachers_dict[teacher_key]
            all_ok = self.extract_and_set_single_teacher_availability_preferences_from_sheet(sheet, teacher_obj,
                                                                                             color_legend_teacher_availability)

            if not all_ok:
                error_count += 1

        if error_count > 0:
            raise Exception(f" {error_count} errors in teacher availability preferences")

        return None

    def _teacher_sanity_checks(self, all_teachers_dict, all_teachers_list, all_classes):
        teachers_to_ignore = []
        # check every valid teacher has availability preferences
        for teacher_obj in all_teachers_list:
            key = teacher_obj["key"]
            availability_preference_table = teacher_obj['availability_preference_table']

            if availability_preference_table is None:
                Logger.warn(
                    f"[{self.log_name}] teacher {Logger.get_teacher_full(teacher_obj)} has no availability plan -> teacher will not be used / removed!")
                teachers_to_ignore.append(teacher_obj)
                continue

            # check if teacher perf classes are known classes
            for day_index, availability_preference_list in enumerate(availability_preference_table):
                for slot_index, slot_obj in enumerate(availability_preference_list):
                    prefilled_class = slot_obj['class_key']
                    if prefilled_class is not None:
                        is_known_class = False
                        for class_obj in all_classes:
                            if prefilled_class == class_obj['key']:
                                is_known_class = True
                                break

                        if not is_known_class:
                            Logger.warn(
                                f"[{self.log_name}] teacher {Logger.get_teacher_full(teacher_obj)} has availability plan entry for unknown class '{prefilled_class}' on day index '{day_index}' slot '{slot_index}' -> slot will be set to skip/not allowed")
                            slot_obj['allowed'] = False
                            slot_obj['class_key'] = None

        for teacher_obj in teachers_to_ignore:
            self._remove_teacher_and_from_class_subjects(teacher_obj, all_teachers_dict, all_teachers_list, all_classes)

        error_count = 0

        # make sure all teachers are still there in the class subject list
        # normally this cannot happen because we remove all invalid/not used teachers (that have no plan) and
        #   their connection to the classes with subjects...
        for class_obj in all_classes:
            class_key = class_obj["key"]
            class_name_single_line = class_obj['name_single_line']

            for subject_obj in class_obj['subjects']:
                subject_name = subject_obj["name"]
                teachers_with_hours = subject_obj['teachers_with_hours']

                for teacher_info in teachers_with_hours:
                    teacher_key = teacher_info['teacher_key']
                    teacher_full_name = teacher_info['teacher_full_name']

                    if teacher_key not in all_teachers_dict:
                        Logger.error(
                            f"[{self.log_name}] teacher '{teacher_key}' [{teacher_full_name}] not found in teacher list for class '{class_key}' [{class_name_single_line}] and subject '{subject_name}'")
                        # raise Exception(msg)
                        error_count += 1

        if error_count > 0:
            # print(f"TODO {error_count} errors in teacher availability preferences")
            raise Exception(f" {error_count} errors in teacher availability preferences")

    # we only use blacklists
    # we set for every slot if it should be filled or not
    def make_teacher_availability_and_prefs_canonical(self, all_teachers_list, color_legend_teacher_availability):
        for teacher_obj in all_teachers_list:
            key = teacher_obj["key"]
            availability_preference_table = teacher_obj['availability_preference_table']

            # we made sure we only have allowed or not allowed colors in availability plans
            # no color -> blacklist
            # allow color -> invert all

            # we set for every slot if it should be filled or not
            is_blacklist = True

            for day_slots_list in availability_preference_table:
                for slot_obj in day_slots_list:
                    if slot_obj["color"] in color_legend_teacher_availability['allowed_bg_colors_set']:
                        is_blacklist = False
                        break
                if not is_blacklist:
                    break

            if is_blacklist:
                # set allowed
                # the colored cells are slots where the teacher is NOT available
                for day_slots_list in availability_preference_table:
                    for slot_obj in day_slots_list:
                        prefilled_class = slot_obj["class_key"]
                        if prefilled_class is not None:
                            slot_obj["allowed"] = False  # already filled -> don't change
                            continue

                        if slot_obj["color"] in color_legend_teacher_availability['not_allowed_bg_colors_set']:
                            slot_obj["allowed"] = False  # should not be filled -> teacher not available here
                        else:
                            slot_obj["allowed"] = True

            else:
                # whitelist ... set all to false except where we have color allowed_bg_color
                # if something is prefilled -> don't change -> not allowed

                # the colored cells are the ONLY slots where the teacher IS available
                for day_slots_list in availability_preference_table:
                    for slot_obj in day_slots_list:
                        prefilled_class = slot_obj["class_key"]
                        if prefilled_class is not None:
                            slot_obj["allowed"] = False  # already filled -> don't change
                            continue

                        if slot_obj["color"] in color_legend_teacher_availability["allowed_bg_colors_set"]:
                            slot_obj["allowed"] = True  # should  be filled -> teacher is available here
                        else:
                            slot_obj["allowed"] = False

    def add_class_key_to_all_classes(self, all_classes, class_key_to_full_name_dict):
        classes_to_remove = []

        for class_obj in all_classes:
            class_name_full = class_obj["name_single_line"]

            for class_key, class_name_infos in class_key_to_full_name_dict.items():
                if class_name_infos['name_full'] == class_name_full:
                    class_obj["key"] = class_key
                    Logger.debug(f"[{self.log_name}] class '{class_name_full}' has key '{class_key}'")
                    break
            if "key" not in class_obj:
                Logger.warn(
                    f"[{self.log_name}]  class '{class_name_full}' has no key in sheet '{MAPPINGS_SHEET_CLASSES}' but has data in hours data excel table, DISCARDING!")
                classes_to_remove.append(class_obj)

        for class_obj in classes_to_remove:
            all_classes.remove(class_obj)

    # see extract_single_teacher_preferences_from_sheet
    def extract_current_single_plan_from_sheet(self, ws_plan, curr_row, curr_col, class_name,
                                               color_legend_teacher_availability):
        curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)

        slots_per_day = 8
        num_days = 5

        dates = []
        # Mo till Fr (the dates)
        for col in range(curr_col + 1, curr_col + num_days + 1):
            date_cell = ws_plan.cell(row=curr_row + 1, column=col)
            dates.append(date_cell.value)

        table = []

        day_index = 0
        for col in range(curr_col + 1, curr_col + num_days + 1):
            day_slots = []
            for row in range(curr_row + 2, curr_row + slots_per_day + 2, 2 if ONLY_USE_BLOCKS_OF_TWO else 1):
                cell = ws_plan.cell(row=row, column=col)

                # cell_color = get_cell_color(cell)
                # just ignore anything in that cell -> we should not fill / process it
                # should_ignore_cell = cell_color == color_legend_teacher_availability["do_not_process_bg_color"]

                day_slots.append({
                    "entry": cell.value,
                    "ignore": False
                })

            should_ignore_day = False
            ignore_day_cell = ws_plan.cell(row=curr_row + slots_per_day + 2, column=col)

            if self.get_cell_color(ignore_day_cell) in color_legend_teacher_availability['not_allowed_bg_colors_set']:
                should_ignore_day = True
                Logger.log(
                    f"[{self.log_name}][Sheet {SHEET_OVERVIEW_PLAN}] ignoring whole day index {day_index}/{num_days} ({dates[day_index]}) because of color below column for class '{class_name}'")

            # set all slots to ignore
            if should_ignore_day:
                for slot_obj in day_slots:
                    slot_obj['ignore'] = True

            table.append(day_slots)
            day_index += 1

        return dates, table

    # extract the current state of the plan from the sheet
    # the task is to fill out ONLY the missing fields
    def extract_current_plan_from_sheet(self, wb_prefs, color_legend_teacher_availability):
        ws_plan = wb_prefs[SHEET_OVERVIEW_PLAN]

        start_col = 2
        start_row = 10
        max_row = 1000

        # key is the class name
        current_plan_dict = {}

        col_increment = 8
        row_increment = 13

        curr_row = start_row
        curr_col = start_col

        all_table_data_dict = {}

        # we check vertically and then horizontally, then again vertically, ...
        is_finished_vertically = False
        is_finished_horizontally = False

        while is_finished_vertically == False:
            is_finished_horizontally = False

            while is_finished_horizontally == False:
                curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)
                class_key = curr_class_name_cell.value

                # TODO sometimes we have invalid class names -> check against dict if we know this class!!
                # e.g. Reserve1, ReserveX, ...

                if class_key is None:
                    is_finished_horizontally = True
                    break
                else:
                    # extract single
                    table_data = self.extract_current_single_plan_from_sheet(ws_plan, curr_row, curr_col, class_key,
                                                                             color_legend_teacher_availability)
                    table_dates = table_data[0]
                    table_column_data = table_data[1]
                    # here is the class name
                    table_start_row = curr_row
                    table_start_col = curr_col

                    all_table_data_dict[class_key] = [table_dates, table_column_data, {
                        "start_row": table_start_row,
                        "start_col": table_start_col
                    }]

                curr_col += col_increment

            curr_col = start_col
            curr_row += row_increment
            curr_class_name_cell = ws_plan.cell(row=curr_row, column=curr_col)

            if curr_class_name_cell.value is None:
                is_finished_vertically = True
                break

        return all_table_data_dict

    def _set_class_plans(self, all_table_data_dict, all_classes):
        # all_table_data_dict[class_key] = [table_dates, table_column_data, {
        #             "start_row": table_start_row,
        #             "start_col": table_start_col
        #         }]

        used_class_keys = []

        for class_obj in all_classes:
            class_key = class_obj["key"]
            if class_key not in all_table_data_dict:
                Logger.warn(f"class key '{class_key}' was not found in table data (Sheet {SHEET_OVERVIEW_PLAN}) but has data in hours data excel table with subjects")
                continue
            used_class_keys.append(class_key)
            class_obj['table_dates'] = all_table_data_dict[class_key]

        # we found plan (schedule) tables for these classes
        all_table_data_class_keys = list(all_table_data_dict.keys())

        difference_class_keys = list(set(all_table_data_class_keys) - set(used_class_keys))

        for class_key in difference_class_keys:
            Logger.warn(f"[{self.log_name}] class key '{class_key}' has table data (Sheet {SHEET_OVERVIEW_PLAN}) but was not found in all classes (hours data excel table with subjects) -> ignoring class")


    def _validate_class_and_teachers(self, all_classes, all_teachers_list, all_teachers_dict):
        Logger.log(f"[{self.log_name}] --- validating class teachers ---")

        error_count = 0

        classes_to_remove = []

        for class_obj in all_classes:
            class_subjects = class_obj["subjects"]

            class_log_name = f"'{class_obj['key']}' ({class_obj['name_single_line']})"

            if len(class_subjects) == 0:
                Logger.warn(
                    f"[{self.log_name}][Validation] class {class_log_name} has no subjects -> will not be processed")
                continue

            class_subjects_to_remove = []

            for class_subject in class_subjects:
                teachers_to_remove = []
                subject_teacher_objs = class_subject["teachers_with_hours"]
                hours_term = class_subject["hours_term"]

                if type(hours_term) == str:
                    Logger.warn(
                        f"[{self.log_name}][Validation] class {class_log_name} has subject '{class_subject['name']}' with hours term '{hours_term}' -> subject will be removed from class")
                    class_subjects_to_remove.append(class_subject)
                    continue

                if hours_term <= 0:
                    Logger.warn(
                        f"[{self.log_name}][Validation] class {class_log_name} has subject '{class_subject['name']}' with hours term '{hours_term}' -> subject will be removed from class")
                    class_subjects_to_remove.append(class_subject)
                    continue

                if len(subject_teacher_objs) == 0:
                    Logger.warn(
                        f"[{self.log_name}][Validation] class {class_log_name} subject '{class_subject['name']}' has no teachers assigned --> subject will be removed from class")
                    class_subjects_to_remove.append(class_subject)
                    # error_count += 1
                    continue

                for subject_teacher in subject_teacher_objs:
                    teacher_key = subject_teacher["teacher_key"]
                    teacher_hours = subject_teacher["hours"]

                    if teacher_key not in all_teachers_dict:
                        Logger.error(
                            f"[{self.log_name}][Validation] class {class_log_name} subject '{class_subject['name']}'  has unknown teacher key '{teacher_key}' assigned --> teacher will be removed from subject")
                        # class_subjects_to_remove.append(class_subject)
                        teachers_to_remove.append(subject_teacher)
                        error_count += 1
                        continue

                for subject_teacher in teachers_to_remove:
                    class_subject["teachers_with_hours"].remove(subject_teacher)

            for class_subject in class_subjects_to_remove:
                class_obj["subjects"].remove(class_subject)

            if len(class_obj["subjects"]) == 0:
                Logger.warn(f"[{self.log_name}][Validation] class {class_log_name} has no subjects left after validation -> will not be processed")
                classes_to_remove.append(class_obj)
                continue

        for class_obj in classes_to_remove:
            class_log_name = f"'{class_obj['key']}' ({class_obj['name_single_line']})"
            all_classes.remove(class_obj)
            Logger.warn(f"[{self.log_name}][Validation] class {class_log_name} has no subjects left after validation -> class will not be processed")

        used_teachers = set()
        teachers_to_remove = []

        # check if every teacher has a subject in any class
        for class_obj in all_classes:
            class_key = class_obj['key']
            subjects_info = class_obj["subjects"]

            for subject_info in subjects_info:
                subject_key = subject_info["name"]
                teachers_with_hours = subject_info["teachers_with_hours"]  # list

                for teachers_with_hour_tuple in teachers_with_hours:
                    teacher_key = teachers_with_hour_tuple['teacher_key']
                    total_required_hours = teachers_with_hour_tuple['hours']
                    used_teachers.add(teacher_key)

        for teacher_obj in all_teachers_list:
            teacher_key = teacher_obj["key"]
            if teacher_key not in used_teachers:
                Logger.warn(f"[{self.log_name}][Validation] teacher {Logger.get_teacher_full(teacher_obj)} not used in any class -> teacher will be removed from all")
                teachers_to_remove.append(teacher_obj)

        for teacher_obj in teachers_to_remove:
            teacher_key = teacher_obj["key"]
            all_teachers_list.remove(teacher_obj)
            del all_teachers_dict[teacher_key]

        if error_count > 0:
            # print(f"TODO {error_count} errors in teacher availability preferences")
            raise Exception(f"[{self.log_name}][Validation] {error_count} errors in teacher availability preferences")


if __name__ == '__main__':
    excel_extractor = ExcelExtractorGesamtuebersicht(
        'example/SJ 25-26_Gesamtübersicht Einsatz Lehrkräfte EA Halle 2025-05-21_3.xlsx',
        "example/Mappings.xlsx",
        'example/07_KW 45_03.11.-07.11.2025_IN.xlsm'
    )
    excel_extractor.read_all()
    excel_extractor.get_max_slots_per_day()
