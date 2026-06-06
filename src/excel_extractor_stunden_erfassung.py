# https://openpyxl.readthedocs.io/en/stable/tutorial.html

import openpyxl

from src.logger import Logger

# e.g. Std-erfassung_KP24A_1. und 2. Sj..xlsx
STD_ERFASSUNG_FILE_PREFIX_LOWER = "std-erfassung_"

SHEET_HOURS_OVERVIEW = 'LehrerStundenübersicht'

ALL_OVERVIEW_NAME = "Gesamtübersicht"

# index of soll year X
# we expect +2 to be the current akkum hours
YEAR_COLUMN_INDICES = [5, 10, 15, 20]
CURR_ACCUM_OFFSET = 2

# TODO can we extract the subject from teacher + class alone?
class ExcelExtractorStundenerfassung:
    def __init__(self, dir_excel_files_stundenerfassung):
        self.log_name = "ExcelStdErfassung"
        self.dir_excel_files_stundenerfassung = dir_excel_files_stundenerfassung
        # the soll index of the year
        # 5 = J
        self.year_columns = YEAR_COLUMN_INDICES
        self.all_soll_data_dict = {}

        self.all_excel_files_stundenerfassungen = self._get_all_std_erfassung_files(
            self.dir_excel_files_stundenerfassung
        )

        self.all_excel_std_files_with_class_obj = []


    def _get_all_std_erfassung_files(self, directory):
        import os
        # the full file name and the class name (everything after [STD_ERFASSUNG_FILE_PREFIX] and before the next "_" or whitespace as tuple)
        excel_files_tuples = []
        for filename in os.listdir(directory):
            # must be a file, not a directory
            if not os.path.isfile(os.path.join(directory, filename)):
                continue
            # must end with .xlsx and contain "Std-erfassung"
            if not filename.endswith(".xlsx"):
                continue

            if not filename.lower().startswith(STD_ERFASSUNG_FILE_PREFIX_LOWER):
                continue

            # get the class name
            file_name_without_prefix = filename[len(STD_ERFASSUNG_FILE_PREFIX_LOWER):]
            next_underscore_index = file_name_without_prefix.find("_")
            if next_underscore_index == -1:
                next_whitespace_index = file_name_without_prefix.find(" ")
                if next_whitespace_index == -1:
                    class_name = file_name_without_prefix.strip()
                else:
                    class_name = file_name_without_prefix[:next_whitespace_index].strip()
            else:
                class_name = file_name_without_prefix[:next_underscore_index].strip()

            excel_full_path = os.path.join(directory, filename)
            Logger.log(f"[{self.log_name}] found Std Erfassung file for class '{class_name}': {excel_full_path}")

            # for some reason the class name/key ends with a '.', ensure that
            if not class_name.endswith("."):
                class_name += "."

            excel_files_tuples.append((excel_full_path, class_name.upper().strip()))

        return excel_files_tuples

    def apply_filter_found_files_with_real_classes(self, all_classes):

        # every class must have a std erfassung file

        used_excel_files_stundenerfassungen_class_names = set()
        error_count = 0
        classes_with_no_excel_file = []

        for class_obj in all_classes:
            class_key = class_obj["key"]

            found_std_erfassung = False

            for excel_file_stundenerfassung_tuple in self.all_excel_files_stundenerfassungen:
                excel_file_path, class_name_from_file_upper = excel_file_stundenerfassung_tuple

                # file name might have wrong casing
                if class_key.upper() == class_name_from_file_upper:
                    found_std_erfassung = True
                    used_excel_files_stundenerfassungen_class_names.add(class_key.upper())
                    Logger.log(f"[{self.log_name}][FILTERING] found matching Std Erfassung file for class '{class_key}': {excel_file_path}")
                    self.all_excel_std_files_with_class_obj.append({
                        "class_obj": class_obj,
                        "excel_file_path": excel_file_path,
                    })
                    break

            if not found_std_erfassung:
                Logger.error(f"[{self.log_name}][FILTERING] no Std Erfassung file found for class '{class_key}' -> this class will be IGNORED!")
                error_count += 1
                classes_with_no_excel_file.append(class_obj)


        excel_files_to_ignore = []
        # now check if all excel files were used
        for excel_file_stundenerfassung_tuple in self.all_excel_files_stundenerfassungen:
            excel_file_path, class_name_from_file_upper = excel_file_stundenerfassung_tuple

            if class_name_from_file_upper.upper() not in used_excel_files_stundenerfassungen_class_names:
                Logger.warn(f"[{self.log_name}][FILTERING] Std Erfassung file for class '{class_name_from_file_upper}' was not used, no matching class found (file: {excel_file_path})!")
                excel_files_to_ignore.append(excel_file_stundenerfassung_tuple)

        for _tuple in excel_files_to_ignore:
            self.all_excel_files_stundenerfassungen.remove(_tuple)

        throw_if_excel_file_not_found = False
        Logger.warn(f"throw_if_excel_file_not_found is set to {throw_if_excel_file_not_found} (if true: every class needs a matching Std Erfassung file else we throw)")
        # print("TODO throw_if_excel_file_not_found!!!!!!!!!!!")

        if error_count > 0:
            if throw_if_excel_file_not_found:
                raise Exception(f"{error_count} classes did not have a matching Std Erfassung file, see previous error messages for details")
            else:
                # remove the classes entirely
                for class_obj in classes_with_no_excel_file:
                    all_classes.remove(class_obj)
                    Logger.warn(f"[{self.log_name}][FILTERING] class '{class_obj['key']}' removed from all_classes because no matching Std Erfassung file was found")

    def read_all(self, ignore_errors_in_std_files):

        some_has_error = False

        if len(self.all_excel_std_files_with_class_obj) == 0:
            raise Exception(f"[{self.log_name}] no 'std erfassung' files found! (maybe call apply_filter_found_files_with_real_classes first?)")

        for obj in self.all_excel_std_files_with_class_obj:
            class_obj = obj["class_obj"]
            excel_file_path = obj["excel_file_path"]

            stundenerfassung_obj, has_error = self.read_single_stundenerfassung(excel_file_path, class_obj, ignore_errors_in_std_files)

            if has_error:
                some_has_error = True

            self.all_soll_data_dict[stundenerfassung_obj["class_key"]] = stundenerfassung_obj["soll_data_array"]

        # we are only allowed to write to "PLAN" sheet
        Logger.log(f"[{self.log_name}][STUNDERFASSUNG] finished reading stundenerfassung excel files")

        if some_has_error and not ignore_errors_in_std_files:
            raise Exception(f"[{self.log_name}][STUNDERFASSUNG] some 'std erfassung' files have errors, fix them")


    def read_single_stundenerfassung(self, excel_files_stundenerfassung_path, class_obj, ignore_errors_in_std_files):
        workbook = openpyxl.load_workbook(excel_files_stundenerfassung_path, read_only=False, data_only=True)
        ws = workbook[SHEET_HOURS_OVERVIEW]

        Logger.log(f"[{self.log_name}][STUNDERFASSUNG] loading 'std erfassung' excel file for class '{class_obj['key']}' from path: {excel_files_stundenerfassung_path}")

        def find_teacher_in_class_subjects(teacher_key, class_obj):
            teacher_with_hours_subject_tuples = []

            for subject_obj in class_obj['subjects']:
                for teacher_with_hours in subject_obj['teachers_with_hours']:
                    if teacher_with_hours['teacher_key'] == teacher_key:
                        teacher_with_hours_subject_tuples.append((teacher_with_hours, subject_obj))

            return teacher_with_hours_subject_tuples

        all_teacher_key_name_tuple_in_class = []

        for subject_obj in class_obj['subjects']:
            for teacher_with_hours in subject_obj['teachers_with_hours']:
                all_teacher_key_name_tuple_in_class.append((teacher_with_hours['teacher_key'], teacher_with_hours['teacher_full_name']))


        found_correct_year_index = False
        correct_year_index = -1

        # class_key = ws.cell(row=1, column=1).value
        class_key = class_obj['key']
        curr_soll_data_array = []
        need_to_save_wb = False
        need_to_log_all_teacher_names_in_class = False

        # in case there are multiple entries for the same teacher and subject --> error
        used_teacher_subject_pairs = set()
        error_count = 0

        for curr_year_index in range(len(YEAR_COLUMN_INDICES)-1, -1, -1):

            # when we can't find the correct subject, output what other subjects the teacher has
            #   (and where in the std erfassung table)
            # entries are also dicts with subject names (so we don't add duplicates in round 2)
            teacher_with_subjects_coords_dict = dict()

            for round in [0,1]:
                Logger.debug(f"[{self.log_name}][Stundenerfassung][{class_key}] --- round {round+1}")
                error_count = 0
                used_teacher_subject_pairs = set()
                # in the first round we read all lines and try to get the data
                #  we need two rounds because we read from top to bottom and
                #  if e.g. a teacher has 2 subjects and we can auto-detect one of it
                #  but the not auto-detectable comes frist we would give an error
                #  however, after we processed the auto-detectable one,
                #  we can auto-process the second one
                is_first_round = True if round == 0 else False
                is_second_round = not is_first_round

                curr_soll_data_array = []
                # we need to determine the correct year index
                # try all years (1-4) and start with the last one
                # if we find at least one entry with != 0 or None entry, then this is the correct year/column

                lfd_nr_col = 1  # A just some number
                subject_col = 27 # AA
                # teacher_key_col = 3 # C
                # name_col = 4  # D
                teacher_subject_pair_col = 3  # C

                start_row = 33
                soll_start_col = self.year_columns[curr_year_index]
                curr_row = start_row - 1

                curr_nr_cell_value = -1


                while curr_nr_cell_value is not None:

                    curr_row += 1
                    curr_nr_cell = ws.cell(row=curr_row, column=lfd_nr_col)

                    # print(curr_nr_cell.value)
                    lfd_nr_value = curr_nr_cell.value
                    curr_nr_cell_value = curr_nr_cell.value

                    if curr_nr_cell_value is None:
                        break

                    soll_cell = ws.cell(row=curr_row, column=soll_start_col)

                    soll_per_week_cell = ws.cell(row=curr_row, column=soll_start_col + 1)
                    soll_per_week_value = soll_per_week_cell.value

                    # soll x. jahr
                    soll_value = soll_cell.value
                    # ist akkum x. jahr
                    ist_akkum_value = ws.cell(row=curr_row, column=soll_start_col + 2).value

                    # this teacher has no hours in this year
                    if soll_value is None or soll_value == "" or soll_value == 0:
                        continue

                    # akt. Soll WS 2. Jahr
                    # this many hours per week are needed to meet the target hours
                    curr_soll_per_week_cell = ws.cell(row=curr_row, column=soll_start_col + 4)
                    curr_soll_per_week_value = curr_soll_per_week_cell.value


                    # only log once
                    if not found_correct_year_index:
                        Logger.log(
                            f"[{self.log_name}][Stundenerfassung][{class_key}] --- found correct year column index: {curr_year_index} (column {soll_cell.column_letter} [index {self.year_columns[correct_year_index]}])")

                    found_correct_year_index = True
                    correct_year_index = curr_year_index

                    if soll_per_week_value == 0 or soll_per_week_value == 0.0:
                        if is_second_round:
                            Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] 'soll per week' is 0 for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(soll_per_week_cell)}' -> this should not happen, please check if the soll per week value is correct, using 0.01 as default")
                        soll_per_week_value = 0.01

                    if type(curr_soll_per_week_value) != float and type(curr_soll_per_week_value) != int and is_second_round:
                        # Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] 'soll per week' is not a number(float) for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(soll_per_week_cell)}', value: {soll_per_week_value} -> this should not happen, please check if the soll per week value is correct")
                        raise Exception(
                            f"[{self.log_name}][Stundenerfassung][{class_key}] 'soll per week' is not a number(float) for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(soll_per_week_cell)}', value: {soll_per_week_value} -> this should not happen, please check if the soll per week value is correct")

                    # teacher_name_value = ws.cell(row=curr_row, column=name_col).value
                    # abkürzung
                    teacher_subject_pair_cell = ws.cell(row=curr_row, column=teacher_subject_pair_col)
                    teacher_subject_pair_value = teacher_subject_pair_cell.value


                    subject_name_cell = ws.cell(row=curr_row, column=subject_col)
                    subject_name_value = subject_name_cell.value

                    if subject_name_value is not None and type(subject_name_value) != str:
                        raise Exception(f"[{self.log_name}][Stundenerfassung][{class_key}] subject name must be a string, cell: {Logger.get_cell_full_coord(subject_name_cell)}, value: {subject_name_value}")


                    # try to extract the teacher from teacher_subject_pair_value value
                    # e.g. Ma(Wind)
                    teacher_start = teacher_subject_pair_value.find("(")
                    teacher_end = teacher_subject_pair_value.rfind(")")
                    teacher_key = teacher_subject_pair_value[teacher_start+1:teacher_end]
                    maybe_subject_name = teacher_subject_pair_value[:teacher_start].strip()

                    if type(teacher_key) != str:
                        raise Exception(
                            f"[{self.log_name}][Stundenerfassung][{class_key}] teacher key must be a string (in teacher + subject cell), cell: {Logger.get_cell_full_coord(teacher_subject_pair_cell)}, value: {teacher_subject_pair_value}")

                    # teacher can have multiple subjects in this class
                    teacher_with_hours_subject_tuples = find_teacher_in_class_subjects(teacher_key, class_obj)

                    if is_first_round:
                        if teacher_key not in teacher_with_subjects_coords_dict:
                            teacher_with_subjects_coords_dict[teacher_key] = dict()

                    correct_subject_obj = None

                    if len(teacher_with_hours_subject_tuples) == 0:
                        # teacher not found in subjects -> error
                        if is_second_round:
                            Logger.error(f"[{self.log_name}][Stundenerfassung][{class_key}] teacher key '{teacher_key}' [Abkürzung: {teacher_subject_pair_value}] not found in class subjects -> PLEASE FIX teacher key at cell: '{Logger.get_cell_full_coord(teacher_subject_pair_cell)} or in overview file [{ALL_OVERVIEW_NAME}]', IGNORING TEACHER")
                        need_to_log_all_teacher_names_in_class = True

                        if is_second_round:
                            error_count += 1
                        continue

                    elif len(teacher_with_hours_subject_tuples) == 1:
                        # this is good, we found the correct teacher and subject
                        correct_teacher_with_hours = teacher_with_hours_subject_tuples[0][0]
                        correct_subject_obj = teacher_with_hours_subject_tuples[0][1]

                        if subject_name_value is None:
                            # subject_name_cell.value = correct_subject_obj['name'].strip() # no need to make it explicit if it's distinct
                            # need_to_save_wb = True
                            subject_name_value = correct_subject_obj['name'].strip()
                        else:
                            # check if correct, if not -> fix
                            if subject_name_value.strip() != correct_subject_obj['name'].strip():
                                Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] subject name '{subject_name_value}' does not match the (correct by teacher and class) subject '{correct_subject_obj['name']}' found for teacher key '{teacher_key}' in class '{class_key}' at cell '{Logger.get_cell_full_coord(teacher_subject_pair_cell)}' -> OVERWRITING FOR YOU")
                                subject_name_cell.value = correct_subject_obj['name']
                                need_to_save_wb = True

                        if correct_subject_obj['name'] not in teacher_with_subjects_coords_dict[teacher_key]:
                            teacher_with_subjects_coords_dict[teacher_key][correct_subject_obj['name']] = {
                                'subject_name': correct_subject_obj['name'],
                                'teacher_subject_pair_cell_coord': Logger.get_cell_full_coord(teacher_subject_pair_cell),
                                'teacher_subject_pair_cell_value': teacher_subject_pair_value,
                            }

                        if is_second_round:
                            Logger.log(f"[{self.log_name}][Stundenerfassung][{class_key}] found correct teacher key '{teacher_key}' with subject '{correct_subject_obj['name']}' for teacher + subject pair value '{teacher_subject_pair_value}' at cell '{Logger.get_cell_full_coord(teacher_subject_pair_cell)}' (from overview file [{ALL_OVERVIEW_NAME}] this class has only this teacher for this subject)  -> using this subject")
                    else:
                        # teacher has multiple subjects in this class
                        # resort to subject name (if any)
                        if subject_name_value is None:

                            # TODO if the other entries for this teacher e.g. Deu & LF4 are empty for this year index
                            # we can identify the correct subject because the other subjects have soll (should) 0 in this year index
                            # we don't try to resolve if there are 0 hours for the other subject (if we have only 2 for a teacher)!
                            #  because normally in the plan there would 0 hours

                            # try to extract the teacher and subject from teacher_subject_pair_value value
                            # e.g. Ma(Wind)
                            # if we cannot extract the teacher or subject -> error
                            # if we can but subject cell was not set -> set it (we don't do it right now, not needed)
                            found_correct_subject_obj = False
                            possible_subjects = []
                            possible_subject_objs = []
                            for i, teacher_subject_tuple in enumerate(teacher_with_hours_subject_tuples):
                                _teacher_with_hours = teacher_subject_tuple[0]
                                _subject_obj = teacher_subject_tuple[1]
                                possible_subjects.append(_subject_obj['name'])
                                possible_subject_objs.append(_subject_obj)

                                if _subject_obj['name'] == maybe_subject_name:
                                    # we found a possible choice (subject) that has exactly this name (in overview file)... take it
                                    # subject_name_cell.value = correct_subject_obj['name'].strip() # no need to make it explicit if it's distinct
                                    # need_to_save_wb = True
                                    correct_subject_obj = _subject_obj
                                    subject_name_value = correct_subject_obj['name'].strip()

                                    # subject comes from the overview for each class -> subjecr should be correct here
                                    # if type(subject_name_value) != str:
                                    #     raise Exception(
                                    #         f"[{self.log_name}][Stundenerfassung][{class_key}] subject name must be a string, cell: {Logger.get_cell_full_coord(subject_name_cell)}, value: {subject_name_value}")

                                    if is_second_round:
                                        Logger.debug(f"[{self.log_name}][Stundenerfassung][{class_key}] found possible subject '{subject_name_value}' for teacher key '{teacher_key}' at cell '{Logger.get_cell_full_coord(teacher_subject_pair_cell)}' based on the exact names in the teacher subject pair value '{teacher_subject_pair_value}' (subject name was taken from overview file) -> using this subject")

                                    found_correct_subject_obj = True
                                    if correct_subject_obj['name'] not in teacher_with_subjects_coords_dict[teacher_key]:
                                        teacher_with_subjects_coords_dict[teacher_key][correct_subject_obj['name']] = {
                                            'subject_name': correct_subject_obj['name'],
                                            'teacher_subject_pair_cell_coord': Logger.get_cell_full_coord(
                                                teacher_subject_pair_cell),
                                            'teacher_subject_pair_cell_value': teacher_subject_pair_value,
                                        }

                            if not found_correct_subject_obj and is_second_round:
                                already_known_subject_names = []
                                # maybe we can reduce the remaining options because we can rule out other subject pairs?
                                # ['Sport', 'LF7'] and we already know/processed the entry for 'LF7' then it must be 'Sport'
                                already_known_subjects = []
                                for subject_name, other_subjects_help in teacher_with_subjects_coords_dict[teacher_key].items():
                                    already_known_subject_names.append(subject_name)
                                    for _subject_obj in possible_subject_objs:
                                        if subject_name == _subject_obj['name']:
                                            # already processed / known
                                            already_known_subjects.append(_subject_obj)
                                            break

                                for obj in already_known_subjects:
                                    possible_subject_objs.remove(obj)

                                if len(possible_subject_objs) == 1:
                                    found_correct_subject_obj = True
                                    correct_subject_obj = possible_subject_objs[0]
                                    subject_name_value = correct_subject_obj['name']
                                    Logger.log(f"[{self.log_name}][Stundenerfassung][{class_key}] teacher key '{teacher_key}' [Abkürzung: {teacher_subject_pair_value}] has multiple subjects in class '{class_key}, but other subjects were already assigned, so we can identify the correct subject, choices were: {possible_subjects} -> auto selected '{subject_name_value}' [because already identified: {already_known_subject_names}]")


                            if not found_correct_subject_obj:
                                other_subjects_help_string = "other subjects from this teacher: "
                                for subject_name, other_subjects_help in teacher_with_subjects_coords_dict[teacher_key].items():
                                    other_subjects_help_string += f"'{other_subjects_help['teacher_subject_pair_cell_value']}' at cell '{other_subjects_help['teacher_subject_pair_cell_coord']}', "

                                if is_second_round:
                                    Logger.error(
                                        f"[{self.log_name}][Stundenerfassung][{class_key}] teacher key '{teacher_key}' [Abkürzung: {teacher_subject_pair_value}] has multiple subjects in class '{class_key}' -> PLEASE set the correct subject name for value '{teacher_subject_pair_value}' at cell: '{Logger.get_cell_full_coord(subject_name_cell)}', the choices are: {possible_subjects}. {other_subjects_help_string}")
                                    error_count += 1
                                continue
                        else:
                            # we have a subject here (explicitly set)
                            for teacher_subject_tuple in teacher_with_hours_subject_tuples:
                                _teacher_with_hours = teacher_subject_tuple[0]
                                _subject_obj = teacher_subject_tuple[1]

                                if _subject_obj['name'].strip() == subject_name_value.strip():
                                    correct_subject_obj = _subject_obj
                                    if correct_subject_obj['name'] not in teacher_with_subjects_coords_dict[teacher_key]:
                                        teacher_with_subjects_coords_dict[teacher_key][correct_subject_obj['name']] = {
                                            'subject_name': correct_subject_obj['name'],
                                            'teacher_subject_pair_cell_coord': Logger.get_cell_full_coord(
                                                teacher_subject_pair_cell),
                                            'teacher_subject_pair_cell_value': teacher_subject_pair_value,
                                        }
                                    break

                            if correct_subject_obj is None:
                                possible_subjects = []
                                for teacher_subject_tuple in teacher_with_hours_subject_tuples:
                                    _teacher_with_hours = teacher_subject_tuple[0]
                                    _subject_obj = teacher_subject_tuple[1]
                                    possible_subjects.append(_subject_obj['name'])

                                other_subjects_help_string = "other subjects from this teacher: "
                                for subject_name, other_subjects_help in teacher_with_subjects_coords_dict[teacher_key].items():
                                    other_subjects_help_string += f"'{other_subjects_help['teacher_subject_pair_cell_value']}' at cell '{other_subjects_help['teacher_subject_pair_cell_coord']}', "

                                if is_second_round:
                                    Logger.error(
                                        f"[{self.log_name}][Stundenerfassung][{class_key}] teacher key '{teacher_key}' [Abkürzung: {teacher_subject_pair_value}] has multiple subjects in class '{class_key}' but none of them match the provided subject name '{subject_name_value}' found in the excel sheet at cell '{Logger.get_cell_full_coord(subject_name_cell)}' -> PLEASE FIX the subject name, the choices are: {possible_subjects}. {other_subjects_help_string}")
                                    error_count += 1
                                continue


                    if teacher_subject_pair_value is None:
                        Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] 'teacher subject pair' is None for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(teacher_subject_pair_cell)}' -> ignoring this entry")
                        continue

                    if soll_value is None:
                        Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] 'soll' is None for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(soll_cell)}' -> ignoring this entry")
                        continue

                    if ist_akkum_value is None:
                        Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] 'ist akkum' is None for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(soll_cell)}' -> ignoring this entry")
                        continue

                    if subject_name_value is None:
                        Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] 'subject name' is None for lfd nr {lfd_nr_value} at cell '{Logger.get_cell_full_coord(subject_name_cell)}' -> ignoring this entry")
                        continue

                    # these should already be strings...
                    teacher_subject_pair = (str(teacher_key),str(subject_name_value))
                    if teacher_subject_pair in used_teacher_subject_pairs:
                        # TODO ENABLE?
                        # raise Exception(f"[{self.log_name}][Stundenerfassung][{class_key}] duplicate entry for teacher key '{teacher_key}' and subject '{subject_name_value}' found in excel sheet at cell '{Logger.get_cell_full_coord(teacher_subject_pair_cell)}' -> this should not happen, please fix the excel sheet, there should be only one entry for each teacher and subject combination")
                        Logger.error(
                            f"[{self.log_name}][Stundenerfassung][{class_key}] duplicate entry for teacher key '{teacher_key}' and subject '{subject_name_value}' found in excel sheet at cell '{Logger.get_cell_full_coord(teacher_subject_pair_cell)}' -> this should not happen, please fix the excel sheet, there should be only one entry for each teacher and subject combination, IGNORING ENTRY")
                        error_count += 1
                        continue

                    used_teacher_subject_pairs.add(teacher_subject_pair)

                    soll_info = {
                        "lfd_nr": lfd_nr_value,
                        # "teacher_name": teacher_name_value,
                        "teacher_subject_pair": teacher_subject_pair_value.strip(),
                        "soll": soll_value,
                        "soll_per_week": soll_per_week_value,
                        "curr_soll_per_week": curr_soll_per_week_value,
                        "ist": ist_akkum_value,
                        "teacher_key": teacher_key.strip(),
                        "subject_name": subject_name_value.strip(),
                    }
                    curr_soll_data_array.append(soll_info)

            if found_correct_year_index:
                break

        if need_to_save_wb:
            workbook.save(excel_files_stundenerfassung_path)

        workbook.close()

        if need_to_log_all_teacher_names_in_class:
            Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] some teacher keys could not be matched to teachers in the class subjects, here are all the teachers in the class for better debugging:")
            for teacher_key_name_tuple in all_teacher_key_name_tuple_in_class:
                Logger.warn(f"[{self.log_name}][Stundenerfassung][{class_key}] - teacher key '{teacher_key_name_tuple[0]}' with name: '{teacher_key_name_tuple[1]}'")

        if not found_correct_year_index:
            Logger.error(f"[{self.log_name}][Stundenerfassung][{class_key}] no correct year column with soll values found for class '{class_key}' -> PLEASE FIX the excel file, no soll values found in any of the expected year columns: {self.year_columns}")
        else:
            Logger.debug(
                f"[{self.log_name}][Stundenerfassung][{class_key}] correct year column index: {correct_year_index} (column {self.year_columns[correct_year_index]}) with {len(curr_soll_data_array)} entries")

        if error_count > 0:
            Logger.error(f"[{self.log_name}][Stundenerfassung][{class_key}] {error_count} errors found while reading stundenerfassung for class '{class_key}', see previous error messages for details")
            # print("TODO enable error")
            # raise Exception(f"[{self.log_name}][Stundenerfassung][{class_key}] {error_count} errors found while reading stundenerfassung for class '{class_key}', see previous error messages for details")

        return {
            "class_key": class_key,
            "correct_year_index": correct_year_index,
            "soll_data_array": curr_soll_data_array,
        }, error_count > 0

    def add_soll_data_to_class_subject_teachers(self, all_classes, all_teachers_list):

        for class_obj in all_classes:
            class_key = class_obj["key"]

            if class_key not in self.all_soll_data_dict:
                Logger.warn(f"class key '{class_key}' not found in stundenerfassung -> skipping")
                continue

            soll_data_array = self.all_soll_data_dict[class_key]

            # TODO proper matching...
            soll_data_to_remove = []

            for soll_data_obj in soll_data_array:
                teacher_name = soll_data_obj['teacher_name']
                soll_teacher_key = soll_data_obj['teacher_key']
                last_part = teacher_name.split(" ")[-1]
                lfd_nr = soll_data_obj['lfd_nr']
                soll_subject_name = soll_data_obj['subject_name']

                teacher_found = False
                for teacher_obj in all_teachers_list:
                    teacher_key = teacher_obj['key']
                    if teacher_key == soll_data_obj['teacher_key']:
                    # last_last_name = teacher_obj['last_name']
                    # if last_last_name == last_part:
                    #     soll_data_obj['teacher_key'] = teacher_obj['key']
                        teacher_found = True
                        # print(teacher_obj['key'])
                        break
                if not teacher_found:
                    Logger.warn(
                        f"teacher '{teacher_name}' from lfd. Nr. {lfd_nr} not found in teacher list -> skipping")
                    soll_data_to_remove.append(soll_data_obj)
                    continue

                # now check subjects
                subject_found = False
                for subject_obj in class_obj['subjects']:
                    class_subject_name = subject_obj['name'].strip()
                    if soll_subject_name == class_subject_name:
                        subject_found = True
                        # add the ist and soll values to the teachers of this subject
                        for teacher_with_hours in subject_obj['teachers_with_hours']:
                            if teacher_with_hours['teacher_key'] == soll_teacher_key:
                                teacher_with_hours['ist'] = soll_data_obj['ist']
                                teacher_with_hours['soll'] = soll_data_obj['soll']
                        break

                if not subject_found:
                    Logger.warn(f"warning: subject '{soll_subject_name}' from lfd. Nr. {lfd_nr} not found in class '{class_key}' -> skipping")
                    soll_data_to_remove.append(soll_data_obj)
                    continue

            for soll_data_obj in soll_data_to_remove:
                soll_data_array.remove(soll_data_obj)

        pass


if __name__ == '__main__':
    excel_extractor = ExcelExtractorStundenerfassung(
        ['example/Std-erfassung_Erz24A_1. und 2. Sj.xlsx'],
    )
    excel_extractor.read_all()
